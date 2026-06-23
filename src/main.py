import sys
import os
import json
import queue as _queue
import threading
from pathlib import Path

try:
    import winsound as _winsound
    _beep_queue: _queue.Queue = _queue.Queue()

    def _beep_worker() -> None:
        """Single daemon thread that serializes all beep requests."""
        while True:
            freq, dur = _beep_queue.get()
            _winsound.Beep(freq, dur)
            _beep_queue.task_done()

    threading.Thread(target=_beep_worker, daemon=True, name="beep-worker").start()

    def _beep(frequency: int, duration_ms: int) -> None:
        """Queue a beep on the shared audio worker thread (non-blocking)."""
        _beep_queue.put((frequency, duration_ms))

except ImportError:
    def _beep(frequency: int, duration_ms: int) -> None:  # type: ignore[misc]
        pass

# Add project root to Python path to find 'shared' module
# This allows imports like 'from shared.stats_manager import StatsManager' to work
project_root = Path(__file__).parent.parent
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QLabel, QPushButton, QVBoxLayout, QWidget, QFileDialog, QStackedWidget,
    QHBoxLayout, QMessageBox, QLineEdit, QComboBox, QDialog, QFormLayout, QDialogButtonBox, QTabWidget,
    QTreeWidget, QTreeWidgetItem, QTableWidget, QTableWidgetItem, QGroupBox, QScrollArea, QInputDialog,
    QProgressDialog
)
from PySide6.QtGui import QAction, QFont, QCloseEvent, QKeySequence
from PySide6.QtCore import QTimer, QSettings, QSize, Qt, QThread, Signal
from datetime import datetime
from openpyxl.styles import PatternFill
import pandas as pd

from logger import get_logger
from profile_manager import ProfileManager, NetworkError, ValidationError
from session_lock_manager import SessionLockManager
from exceptions import SessionLockedError, StaleLockError
from session_selector import SessionSelectorDialog
from print_dialog import PrintDialog
from packer_mode_widget import PackerModeWidget
from packer_logic import PackerLogic, REQUIRED_COLUMNS
from session_manager import SessionManager
from shared.stats_manager import StatsManager
from shared.worker_manager import WorkerManager
from sku_mapping_dialog import SKUMappingDialog
from session_history_manager import SessionHistoryManager
from session_browser.session_browser_widget import SessionBrowserWidget
from session_registry_manager import SessionRegistryManager
from worker_selection_dialog import WorkerSelectionDialog
from theme import load_saved_theme, toggle_theme

logger = get_logger(__name__)

DEFAULT_CONFIG_PATH = "config.ini"

def find_latest_session_dir(base_dir: str = ".") -> str | None:
    """
    Finds the most recent, valid, and incomplete session directory.

    A session is considered valid and incomplete if it is a directory matching
    the session name pattern and contains a `session_info.json` file.

    Args:
        base_dir (str): The directory to search for session folders.

    Returns:
        str | None: The path to the latest valid session directory, or None if
                    none is found.
    """
    session_pattern = "OrdersFulfillment_"
    session_dirs = []
    try:
        session_dirs = [d for d in os.listdir(base_dir) if os.path.isdir(os.path.join(base_dir, d)) and d.startswith(session_pattern)]
    except FileNotFoundError:
        return None # No base directory yet

    valid_sessions = []
    for dirname in session_dirs:
        dirpath = os.path.join(base_dir, dirname)
        # A valid, incomplete session must have a session_info file.
        if os.path.exists(os.path.join(dirpath, "session_info.json")):
            valid_sessions.append(dirpath)

    if not valid_sessions:
        return None

    # Return the one that was most recently modified
    latest_session_dir = max(valid_sessions, key=os.path.getmtime)
    return latest_session_dir


class SessionStartWorker(QThread):
    """
    Background worker for the slow I/O steps when starting a session.

    Performs PackerLogic construction (reads packer_config + packing_state from server)
    and load_packing_list_json (reads + parses the packing list JSON) off the UI thread.

    Lock acquisition and heartbeat setup remain on the main thread because stale-lock
    handling requires a QMessageBox interaction.

    Usage (blocking-with-progress pattern):
        worker = SessionStartWorker(client_id, profile_manager, work_dir, packing_list_path)
        worker.start()
        while not worker.wait(50):
            QApplication.processEvents()
        if worker.error:
            raise worker.error
        self.logic = worker.logic
    """

    def __init__(self, client_id, profile_manager, work_dir, packing_list_path, parent=None):
        super().__init__(parent)
        self._client_id = client_id
        self._profile_manager = profile_manager
        self._work_dir = work_dir
        self._packing_list_path = packing_list_path
        # Results (read by main thread after wait())
        self.logic = None
        self.order_count = 0
        self.list_name = ""
        self.error = None  # Exception instance if failed

    def run(self) -> None:
        try:
            from packer_logic import PackerLogic
            logic = PackerLogic(
                client_id=self._client_id,
                profile_manager=self._profile_manager,
                work_dir=str(self._work_dir),
            )
            order_count, list_name = logic.load_packing_list_json(str(self._packing_list_path))
            # Move Qt object ownership back to the main thread
            logic.moveToThread(QApplication.instance().thread())
            self.logic = logic
            self.order_count = order_count
            self.list_name = list_name
        except Exception as exc:
            self.error = exc


class SessionEndWorker(QThread):
    """
    Background worker for the slow server-write operations at session end.

    Accepts a single callable (write_fn) that captures all necessary context
    via closure, keeping this class generic and the caller readable.

    Usage (blocking-with-progress pattern):
        worker = SessionEndWorker(lambda: _do_all_slow_writes())
        worker.start()
        while not worker.wait(50):
            QApplication.processEvents()
        if worker.error:
            logger.error(...)
        # Proceed with lock release / UI reset
    """

    def __init__(self, write_fn, parent=None):
        super().__init__(parent)
        self._write_fn = write_fn
        self.error = None

    def run(self) -> None:
        try:
            self._write_fn()
        except Exception as exc:
            logger.error(f"SessionEndWorker: unexpected error: {exc}", exc_info=True)
            self.error = exc


class MainWindow(QMainWindow):
    """
    The main application window, acting as the central orchestrator.

    This class initializes the UI, manages application state, and connects UI
    events to the backend logic. It handles the overall workflow, including
    session management, data loading, and switching between different views.

    Attributes:
        session_manager (SessionManager): Manages the lifecycle of packing sessions.
        logic (PackerLogic | None): The core business logic for the current session.
        stats_manager (StatisticsManager): Manages persistent application statistics.
        session_widget (QWidget): The main widget for the session view.
        packer_mode_widget (PackerModeWidget): The widget for the packer mode view.
        stacked_widget (QStackedWidget): Manages switching between views.
        orders_table (QTableView): The table displaying the list of orders.
        table_model (OrderTableModel): The model for the orders table.
        proxy_model (CustomFilterProxyModel): The proxy model for filtering the table.
    """
    def __init__(self, skip_worker_selection: bool = False, config_path: str = DEFAULT_CONFIG_PATH):
        """Initialize the MainWindow, sets up UI, and loads initial state.

        Args:
            skip_worker_selection: If True, skip worker selection dialog (for tests)
            config_path: Path to the configuration file (default: config.ini).
                Use a dev config (e.g. config.dev.ini) to point at a local mock server.
        """
        super().__init__()
        self.setWindowTitle("Packer's Assistant")
        self.resize(1024, 768)

        logger.info("Initializing MainWindow")

        # Detect if running in test mode
        self._is_test_mode = skip_worker_selection or 'pytest' in sys.modules

        # Initialize ProfileManager (may raise NetworkError)
        try:
            self.profile_manager = ProfileManager(config_path)
            logger.info("ProfileManager initialized successfully")
        except NetworkError as e:
            logger.error(f"Failed to initialize ProfileManager: {e}")
            QMessageBox.critical(
                self,
                "Network Error",
                f"Cannot connect to file server:\n\n{e}\n\n"
                f"Please check your network connection and try again."
            )
            sys.exit(1)
        except Exception as e:
            logger.error(f"Unexpected error initializing ProfileManager: {e}", exc_info=True)
            QMessageBox.critical(self, "Error", f"Failed to initialize application:\n\n{e}")
            sys.exit(1)

        # Initialize SessionLockManager
        self.lock_manager = SessionLockManager(self.profile_manager)
        logger.info("SessionLockManager initialized successfully")

        # Initialize WorkerManager
        base_path = self.profile_manager.base_path
        self.worker_manager = WorkerManager(str(base_path))
        logger.info("WorkerManager initialized successfully")

        # Initialize SessionHistoryManager
        self.session_history_manager = SessionHistoryManager(self.profile_manager)
        logger.info("SessionHistoryManager initialized successfully")

        # Initialize SessionRegistryManager (per-client index for fast browser loading)
        self.registry_manager = SessionRegistryManager(self.profile_manager)
        logger.info("SessionRegistryManager initialized successfully")

        # Read scan simulator mode from config (enabled in development / no physical scanner)
        self._sim_mode = self.profile_manager.config.getboolean(
            'General', 'ScanSimulatorMode', fallback=False
        )
        if self._sim_mode:
            logger.info("Scan Simulator Mode enabled (dev/test environment)")

        # Worker state
        self.current_worker_id = None
        self.current_worker_name = None

        # Current client state
        self.current_client_id = None
        self.session_manager = None  # Will be instantiated per client
        self.logic = None  # Will be instantiated per session

        # Shopify session state (new workflow)
        self.current_session_path = None  # Path to current Shopify session
        self.current_packing_list = None  # Name of selected packing list
        self.current_work_dir = None      # Work directory for packing results
        self.packing_data = None          # Loaded packing list data

        # Phase 1.4: Unified StatsManager for integration with Shopify Tool statistics
        # Records packing statistics to shared Stats/global_stats.json on file server
        # Used for:
        # 1. Historical analytics and performance tracking across both tools
        # 2. Integration with Shopify Tool (shared statistics file)
        # 3. Warehouse operation audit trail and worker performance metrics
        # 4. Per-client analytics and reporting
        # Note: Called once per session (at completion) by design - records session totals
        self.stats_manager = StatsManager(base_path=str(base_path))

        # Settings for remembering last client
        self.settings = QSettings("PackingTool", "ClientSelection")

        # Show worker selection BEFORE main window initialization (skip in test mode)
        if not self._is_test_mode:
            if not self._select_worker():
                # User cancelled - exit app
                logger.info("Worker selection cancelled - exiting application")
                sys.exit(0)
        else:
            # Test mode - use dummy worker
            self.current_worker_id = "test_worker_001"
            self.current_worker_name = "Test Worker"
            logger.info(f"Test mode: Using dummy worker {self.current_worker_name}")

        self._init_ui()

        # Load available clients and restore last selected
        self.load_available_clients()

        logger.info("MainWindow initialized successfully")

    def _init_ui(self):
        """Initialize all user interface components and layouts."""
        self.session_widget = QWidget()
        main_layout = QVBoxLayout(self.session_widget)

        # (no inline stylesheet — global QSS + QPalette handle all colors and fonts)

        # Set minimum window size
        self.setMinimumSize(900, 600)

        # ====================================================================
        # CLIENT SELECTION (NEW)
        # ====================================================================
        client_selection_widget = QWidget()
        client_selection_widget.setObjectName("ClientSelection")
        client_selection_widget.setStyleSheet(
            "QWidget#ClientSelection { border-bottom: 1px solid palette(mid); padding-bottom: 4px; }"
        )
        client_selection_layout = QHBoxLayout(client_selection_widget)
        client_selection_layout.setContentsMargins(6, 4, 6, 4)
        main_layout.addWidget(client_selection_widget)

        client_label = QLabel("Client:")
        client_label.setStyleSheet("font-size: 12pt; font-weight: bold;")
        client_selection_layout.addWidget(client_label)

        self.client_combo = QComboBox()
        self.client_combo.setMinimumWidth(250)
        self.client_combo.setSizeAdjustPolicy(QComboBox.AdjustToContents)
        self.client_combo.currentIndexChanged.connect(self.on_client_changed)
        client_selection_layout.addWidget(self.client_combo)

        client_selection_layout.addStretch()

        # "+ New Client" button kept as hidden helper — accessible via logic only
        self.new_client_button = QPushButton("+ New Client")
        self.new_client_button.clicked.connect(self.create_new_client)
        self.new_client_button.setVisible(False)  # hidden from main UI

        # (control panel removed — all actions are in the toolbar)

        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Search by Order Number, SKU, or Status...")
        self.search_input.textChanged.connect(self._filter_orders)
        main_layout.addWidget(self.search_input)

        # Create tab widget for session views (Packing and Statistics)
        self.session_tabs = QTabWidget()

        # Tab 1: Packing View with expandable tree
        packing_tab = QWidget()
        packing_layout = QVBoxLayout(packing_tab)
        packing_layout.setContentsMargins(0, 0, 0, 0)

        self._setup_order_tree()
        packing_layout.addWidget(self.order_tree)

        self.session_tabs.addTab(packing_tab, "Packing")

        # Tab 2: Statistics View
        stats_tab = QWidget()
        stats_layout = QVBoxLayout(stats_tab)
        stats_layout.setContentsMargins(0, 0, 0, 0)
        self._setup_statistics_tab(stats_layout)
        self.session_tabs.addTab(stats_tab, "Statistics")

        main_layout.addWidget(self.session_tabs)

        self.status_label = QLabel("Start a new session to load a packing list.")
        self.status_label.setObjectName("status_msg_label")
        self.status_label.setStyleSheet(
            "QLabel#status_msg_label { "
            "border-top: 1px solid palette(mid); "
            "padding: 4px 6px; "
            "font-size: 10pt; "
            "color: #aaaaaa; "
            "}"
        )
        main_layout.addWidget(self.status_label)

        self.packer_mode_widget = PackerModeWidget(sim_mode=self._sim_mode)
        self.packer_mode_widget.barcode_scanned.connect(self.on_scanner_input)
        self.packer_mode_widget.exit_packing_mode.connect(self.switch_to_session_view)
        self.packer_mode_widget.skip_order_requested.connect(self._on_skip_order)
        self.packer_mode_widget.cancel_item_requested.connect(self._on_cancel_item)
        self.packer_mode_widget.force_confirm_requested.connect(self._on_force_confirm)
        self.packer_mode_widget.map_sku_requested.connect(self._on_map_sku_from_packer)
        self.packer_mode_widget.extra_confirmed.connect(self._on_extra_confirmed)
        self.packer_mode_widget.extra_removed.connect(self._on_extra_removed)

        # Stacked widget to switch between session view and packer mode
        self.stacked_widget = QStackedWidget()
        self.stacked_widget.addWidget(self.session_widget)
        self.stacked_widget.addWidget(self.packer_mode_widget)
        self.setCentralWidget(self.stacked_widget)

        # Create menu bar and toolbar
        self._init_menu_bar()
        self._init_toolbar()
        self._init_status_bar()

    def _init_menu_bar(self):
        """Initialize the menu bar with organized actions."""
        menubar = self.menuBar()

        # File menu
        file_menu = menubar.addMenu("&File")

        exit_action = QAction("Exit", self)
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)

        # Session menu
        session_menu = menubar.addMenu("&Session")

        shopify_session_action = QAction("Open Shopify Session...", self)
        shopify_session_action.setShortcut(QKeySequence("Ctrl+O"))
        shopify_session_action.triggered.connect(self.open_shopify_session)
        session_menu.addAction(shopify_session_action)

        browse_action = QAction("Session Browser...", self)
        browse_action.setShortcut(QKeySequence("Ctrl+B"))
        browse_action.triggered.connect(self.open_session_browser)
        session_menu.addAction(browse_action)

        session_menu.addSeparator()

        end_action = QAction("End Current Session", self)
        end_action.setShortcut(QKeySequence("Ctrl+E"))
        end_action.triggered.connect(self.end_session)
        session_menu.addAction(end_action)

        # Settings menu
        settings_menu = menubar.addMenu("&Settings")

        worker_action = QAction("Select Worker...", self)
        worker_action.triggered.connect(self._select_worker)
        settings_menu.addAction(worker_action)

        sku_mapping_action = QAction("SKU Mappings...", self)
        sku_mapping_action.triggered.connect(self.open_sku_mapping_dialog)
        settings_menu.addAction(sku_mapping_action)

        settings_menu.addSeparator()

        theme_action = QAction("Toggle Dark/Light Theme", self)
        theme_action.triggered.connect(self._toggle_theme)
        settings_menu.addAction(theme_action)

    def _init_toolbar(self):
        """Create toolbar with all session actions."""
        from PySide6.QtWidgets import QToolBar

        toolbar = QToolBar()
        toolbar.setMovable(False)
        toolbar.setIconSize(QSize(24, 24))

        # Session start actions (primary blue — main entry points)
        shopify_btn = QPushButton("Shopify Session")
        shopify_btn.clicked.connect(self.open_shopify_session)
        shopify_btn.setToolTip("Open a Shopify packing session")
        shopify_btn.setProperty("primary", "true")
        toolbar.addWidget(shopify_btn)

        browser_btn = QPushButton("Session Browser")
        browser_btn.clicked.connect(self.open_session_browser)
        browser_btn.setToolTip("Browse active, completed, and available sessions")
        browser_btn.setProperty("primary", "true")
        toolbar.addWidget(browser_btn)

        toolbar.addSeparator()

        # Current session info label
        self.session_info_label = QLabel("No active session")
        self.session_info_label.setObjectName("session_info_label")
        font = QFont()
        font.setPointSize(10)
        font.setBold(True)
        self.session_info_label.setFont(font)
        toolbar.addWidget(self.session_info_label)

        toolbar.addSeparator()

        # Packing mode button
        self.packer_mode_button = QPushButton("Start Packing")
        self.packer_mode_button.setEnabled(False)
        self.packer_mode_button.clicked.connect(self.switch_to_packer_mode)
        self.packer_mode_button.setToolTip("Switch to barcode scanning / packer mode")
        self.packer_mode_button.setProperty("primary", "true")
        toolbar.addWidget(self.packer_mode_button)

        # SKU mapping button
        self.sku_mapping_button = QPushButton("SKU Mapping")
        self.sku_mapping_button.clicked.connect(self.open_sku_mapping_dialog)
        self.sku_mapping_button.setToolTip("Manage barcode to SKU mappings")
        toolbar.addWidget(self.sku_mapping_button)

        toolbar.addSeparator()

        # End session button
        self.toolbar_end_btn = QPushButton("End Session")
        self.toolbar_end_btn.setObjectName("danger")
        self.toolbar_end_btn.clicked.connect(self.end_session)
        self.toolbar_end_btn.setEnabled(False)
        self.toolbar_end_btn.setToolTip("End the current packing session")
        toolbar.addWidget(self.toolbar_end_btn)

        self.addToolBar(toolbar)

    def _init_status_bar(self):
        """Set up the bottom status bar with permanent worker label."""
        status_bar = self.statusBar()
        self.sb_worker_label = QLabel(f"Worker: {self.current_worker_name or 'None'}")
        self.sb_worker_label.setObjectName("worker_label")
        status_bar.addPermanentWidget(self.sb_worker_label)

    def _setup_order_tree(self):
        """Setup expandable order tree view."""
        self.order_tree = QTreeWidget()
        self.order_tree.setHeaderLabels([
            "Order / Item", "Product", "Quantity", "Status", "Courier"
        ])

        # Column widths (interactive, with sensible defaults)
        from PySide6.QtWidgets import QHeaderView
        self.order_tree.setColumnWidth(0, 180)  # Order/SKU
        self.order_tree.setColumnWidth(2, 80)   # Quantity
        self.order_tree.setColumnWidth(3, 110)  # Status
        self.order_tree.setColumnWidth(4, 130)  # Courier
        self.order_tree.header().setSectionResizeMode(1, QHeaderView.Stretch)  # Product stretches

        # Styling
        font = QFont()
        font.setPointSize(11)
        self.order_tree.setFont(font)
        self.order_tree.setAlternatingRowColors(True)
        self.order_tree.setUniformRowHeights(False)
        self.order_tree.setItemsExpandable(True)
        self.order_tree.setRootIsDecorated(True)

        # Row height and styling
        self.order_tree.setStyleSheet("""
            QTreeWidget::item {
                height: 30px;
                padding: 5px;
            }
        """)

    def _populate_order_tree(self):
        """Populate tree with orders and items."""
        self.order_tree.clear()

        if not self.logic or not hasattr(self.logic, 'processed_df') or self.logic.processed_df is None:
            return

        # Group by order number
        grouped = self.logic.processed_df.groupby('Order_Number')

        # Get completed and in-progress orders
        completed_orders = self.logic.session_packing_state.get('completed_orders', [])
        in_progress_orders = self.logic.session_packing_state.get('in_progress', {})

        for order_num, order_items in grouped:
            items_df = order_items
            total_items = len(items_df)

            # Check order status
            is_completed = order_num in completed_orders

            # Count scanned items
            scanned_count = 0
            if order_num in in_progress_orders:
                order_state = in_progress_orders[order_num]
                # Handle both list format (current) and dict format (legacy)
                if isinstance(order_state, list):
                    scanned_count = sum(1 for s in order_state if s.get('packed', 0) >= s.get('required', 1))
                else:
                    scanned_count = sum(1 for s in order_state.values() if s.get('packed', 0) >= s.get('required', 1))
            elif is_completed:
                scanned_count = total_items

            # Order status
            if is_completed:
                status_text = "✅ Completed"
                status_icon = "✅"
            else:
                status_text = f"⏳ {scanned_count}/{total_items} items"
                status_icon = "⏳"

            # Courier
            courier = items_df.iloc[0].get('Courier', 'N/A') if 'Courier' in items_df.columns else 'N/A'

            # Create top-level order item
            order_item = QTreeWidgetItem([
                f"{order_num}",
                f"{total_items} items",
                "",
                status_text,
                courier
            ])

            # Bold font for order
            font = QFont()
            font.setBold(True)
            font.setPointSize(11)
            for col in range(5):
                order_item.setFont(col, font)

            # Add child items (SKUs) - OPTIMIZED: replaced iterrows() with itertuples()
            # itertuples() is 5-10x faster than iterrows() for DataFrame iteration
            for row_tuple in items_df.itertuples(index=False):
                # Access by column index from tuple
                sku = getattr(row_tuple, 'SKU', 'Unknown') if hasattr(row_tuple, 'SKU') else 'Unknown'
                product = getattr(row_tuple, 'Product_Name', 'Unknown') if hasattr(row_tuple, 'Product_Name') else 'Unknown'
                qty = getattr(row_tuple, 'Quantity', 1) if hasattr(row_tuple, 'Quantity') else 1

                # Check if scanned
                scanned_qty = 0
                if order_num in in_progress_orders:
                    order_state = in_progress_orders[order_num]
                    # Find this SKU in the order state
                    if isinstance(order_state, list):
                        for item_state in order_state:
                            # CRITICAL FIX: Validate item_state is dict before calling .get()
                            if not isinstance(item_state, dict):
                                logger.warning(f"Skipping invalid item_state in {order_num}: {type(item_state).__name__}")
                                continue

                            if item_state.get('original_sku') == sku:
                                scanned_qty = item_state.get('packed', 0)
                                break
                    else:
                        # Legacy dict format
                        for item_state in order_state.values():
                            # CRITICAL FIX: Validate item_state is dict before calling .get()
                            if not isinstance(item_state, dict):
                                logger.warning(f"Skipping invalid item_state in {order_num}: {type(item_state).__name__}")
                                continue

                            if item_state.get('original_sku') == sku:
                                scanned_qty = item_state.get('packed', 0)
                                break
                elif is_completed:
                    try:
                        scanned_qty = int(qty)
                    except:
                        scanned_qty = 1

                try:
                    qty_int = int(qty)
                except:
                    qty_int = 1

                if scanned_qty >= qty_int:
                    item_status = "✅ Scanned"
                else:
                    item_status = f"⏳ Pending ({scanned_qty}/{qty_int})"

                # Create child item
                child_item = QTreeWidgetItem([
                    f"  {sku}",
                    product,
                    str(qty),
                    item_status,
                    ""
                ])

                # Normal font for items
                item_font = QFont()
                item_font.setPointSize(10)
                for col in range(5):
                    child_item.setFont(col, item_font)

                order_item.addChild(child_item)

            self.order_tree.addTopLevelItem(order_item)

            # Expand completed orders, collapse pending
            if is_completed:
                order_item.setExpanded(False)  # Keep compact
            else:
                order_item.setExpanded(True)   # Show current work

    def _filter_orders(self, text: str):
        """Filter tree items by search text."""
        if not hasattr(self, 'order_tree'):
            return

        if not text:
            # Show all
            for i in range(self.order_tree.topLevelItemCount()):
                self.order_tree.topLevelItem(i).setHidden(False)
            return

        text = text.lower()

        for i in range(self.order_tree.topLevelItemCount()):
            order_item = self.order_tree.topLevelItem(i)
            order_text = order_item.text(0).lower()

            # Check if order matches
            order_match = text in order_text

            # Check if any child (SKU) matches
            child_match = False
            for j in range(order_item.childCount()):
                child = order_item.child(j)
                child_text = f"{child.text(0)} {child.text(1)}".lower()
                if text in child_text:
                    child_match = True
                    break

            # Show if order or child matches
            order_item.setHidden(not (order_match or child_match))

    def _setup_statistics_tab(self, layout):
        """Create statistics overview tab."""

        # Scroll area for stats
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout(scroll_widget)

        # --- Session Totals (horizontal compact cards) ---
        totals_group = QGroupBox("Session Totals")
        totals_row = QHBoxLayout()
        totals_row.setSpacing(16)

        bold_font = QFont()
        bold_font.setPointSize(18)
        bold_font.setBold(True)
        label_font = QFont()
        label_font.setPointSize(9)

        def _make_stat_card(title: str) -> QLabel:
            """Create a vertical mini-card: big value on top, small label below."""
            card = QWidget()
            card_layout = QVBoxLayout(card)
            card_layout.setContentsMargins(12, 8, 12, 8)
            card_layout.setSpacing(2)
            card.setStyleSheet("QWidget { border: 1px solid #2a2a2a; border-radius: 4px; }")
            value_lbl = QLabel("0")
            value_lbl.setFont(bold_font)
            value_lbl.setAlignment(Qt.AlignCenter)
            title_lbl = QLabel(title)
            title_lbl.setFont(label_font)
            title_lbl.setAlignment(Qt.AlignCenter)
            title_lbl.setStyleSheet("color: #888888; border: none;")
            card_layout.addWidget(value_lbl)
            card_layout.addWidget(title_lbl)
            totals_row.addWidget(card)
            return value_lbl

        self.stats_total_orders = _make_stat_card("Orders")
        self.stats_completed_orders = _make_stat_card("Completed")
        self.stats_total_items = _make_stat_card("Items")
        self.stats_unique_skus = _make_stat_card("Unique SKUs")
        self.stats_progress_pct = _make_stat_card("Progress")
        totals_row.addStretch()

        totals_group.setLayout(totals_row)
        scroll_layout.addWidget(totals_group)

        # --- By Courier ---
        courier_group = QGroupBox("By Courier")
        courier_layout = QHBoxLayout()
        self.courier_stats_widget = QWidget()
        self.courier_stats_layout = QHBoxLayout(self.courier_stats_widget)
        self.courier_stats_layout.setSpacing(16)
        courier_layout.addWidget(self.courier_stats_widget)
        courier_layout.addStretch()
        courier_group.setLayout(courier_layout)
        scroll_layout.addWidget(courier_group)

        # --- SKU Summary Table ---
        sku_group = QGroupBox("SKU Summary")
        sku_layout = QVBoxLayout()

        self.sku_table = QTableWidget()
        self.sku_table.setColumnCount(4)
        self.sku_table.setHorizontalHeaderLabels(["SKU", "Product", "Total Qty", "Status"])
        self.sku_table.horizontalHeader().setStretchLastSection(True)
        self.sku_table.setAlternatingRowColors(True)
        self.sku_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.sku_table.setSelectionMode(QTableWidget.NoSelection)

        sku_layout.addWidget(self.sku_table)
        sku_group.setLayout(sku_layout)
        scroll_layout.addWidget(sku_group)

        scroll.setWidget(scroll_widget)
        layout.addWidget(scroll)

    def _update_statistics(self):
        """Refresh statistics tab with current data."""
        if not self.logic or not hasattr(self.logic, 'processed_df') or self.logic.processed_df is None:
            return

        df = self.logic.processed_df

        # Session totals
        unique_orders = df['Order_Number'].unique()
        total_orders = len(unique_orders)
        completed_orders_list = self.logic.session_packing_state.get('completed_orders', [])
        completed_orders = len(completed_orders_list)
        total_items = len(df)
        unique_skus = df['SKU'].nunique()
        progress_pct = int((completed_orders / total_orders * 100)) if total_orders > 0 else 0

        self.stats_total_orders.setText(str(total_orders))
        self.stats_completed_orders.setText(str(completed_orders))
        self.stats_total_items.setText(str(total_items))
        self.stats_unique_skus.setText(str(unique_skus))
        self.stats_progress_pct.setText(f"{progress_pct}%")

        # By Courier
        # Clear existing
        for i in reversed(range(self.courier_stats_layout.count())):
            widget = self.courier_stats_layout.itemAt(i).widget()
            if widget:
                widget.setParent(None)

        if 'Courier' in df.columns:
            courier_stats = df.groupby('Courier').agg({
                'Order_Number': 'nunique',
                'Quantity': lambda x: pd.to_numeric(x, errors='coerce').sum()
            }).reset_index()

            # OPTIMIZED: replaced iterrows() with itertuples() for 5-10x speedup
            card_bold_font = QFont()
            card_bold_font.setPointSize(18)
            card_bold_font.setBold(True)
            card_label_font = QFont()
            card_label_font.setPointSize(9)
            for row_tuple in courier_stats.itertuples(index=False):
                courier = row_tuple.Courier
                orders = row_tuple.Order_Number
                items = int(row_tuple.Quantity) if pd.notna(row_tuple.Quantity) else 0
                card = QWidget()
                card_layout = QVBoxLayout(card)
                card_layout.setContentsMargins(12, 8, 12, 8)
                card_layout.setSpacing(2)
                card.setObjectName("courier_card")
                card.setStyleSheet("#courier_card { border: 1px solid #2a2a2a; border-radius: 4px; }")
                value_lbl = QLabel(str(orders))
                value_lbl.setFont(card_bold_font)
                value_lbl.setAlignment(Qt.AlignCenter)
                courier_lbl = QLabel(courier)
                courier_lbl.setFont(card_label_font)
                courier_lbl.setAlignment(Qt.AlignCenter)
                courier_lbl.setStyleSheet("color: #888888; border: none;")
                items_lbl = QLabel(f"{items} items")
                items_lbl.setFont(card_label_font)
                items_lbl.setAlignment(Qt.AlignCenter)
                items_lbl.setStyleSheet("color: #666666; border: none;")
                card_layout.addWidget(value_lbl)
                card_layout.addWidget(courier_lbl)
                card_layout.addWidget(items_lbl)
                self.courier_stats_layout.addWidget(card)

        # SKU Summary
        sku_summary = df.groupby(['SKU', 'Product_Name']).agg({
            'Quantity': lambda x: pd.to_numeric(x, errors='coerce').sum()
        }).reset_index()

        self.sku_table.setRowCount(len(sku_summary))

        # Get scanned items tracking
        in_progress_orders = self.logic.session_packing_state.get('in_progress', {})
        scanned_by_sku = {}

        # Count scanned quantities per SKU
        for order_state in in_progress_orders.values():
            if isinstance(order_state, list):
                for item_state in order_state:
                    # CRITICAL FIX: Validate item_state is dict before calling .get()
                    if not isinstance(item_state, dict):
                        logger.warning(f"Skipping invalid item_state (not a dict): {type(item_state).__name__}")
                        continue

                    sku = item_state.get('original_sku')
                    packed = item_state.get('packed', 0)
                    if sku:
                        scanned_by_sku[sku] = scanned_by_sku.get(sku, 0) + packed
            else:
                for item_state in order_state.values():
                    # CRITICAL FIX: Validate item_state is dict before calling .get()
                    if not isinstance(item_state, dict):
                        logger.warning(f"Skipping invalid item_state (not a dict): {type(item_state).__name__}")
                        continue

                    sku = item_state.get('original_sku')
                    packed = item_state.get('packed', 0)
                    if sku:
                        scanned_by_sku[sku] = scanned_by_sku.get(sku, 0) + packed

        # Add completed orders to scanned count
        # OPTIMIZED: replaced nested loops + iterrows() with vectorized groupby
        # This reduces O(n*m) iteration to O(n) vectorized operation
        if completed_orders_list:
            completed_items_df = df[df['Order_Number'].isin(completed_orders_list)]
            if not completed_items_df.empty:
                # Group by SKU and sum quantities (vectorized)
                completed_by_sku = completed_items_df.groupby('SKU')['Quantity'].apply(
                    lambda x: pd.to_numeric(x, errors='coerce').sum()
                ).fillna(0).astype(int)

                # Add to scanned_by_sku dict
                for sku, qty in completed_by_sku.items():
                    scanned_by_sku[sku] = scanned_by_sku.get(sku, 0) + qty

        # OPTIMIZED: replaced iterrows() with enumerate(itertuples()) for 5-10x speedup
        for idx, row_tuple in enumerate(sku_summary.itertuples(index=False)):
            sku = row_tuple.SKU
            product = row_tuple.Product_Name
            qty = row_tuple.Quantity
            qty_int = int(qty) if pd.notna(qty) else 0

            # Check if fully scanned
            scanned = scanned_by_sku.get(sku, 0)
            if scanned >= qty_int:
                status = "✅ Complete"
            else:
                status = f"⏳ {scanned}/{qty_int}"

            self.sku_table.setItem(idx, 0, QTableWidgetItem(sku))
            self.sku_table.setItem(idx, 1, QTableWidgetItem(product))
            self.sku_table.setItem(idx, 2, QTableWidgetItem(str(qty_int)))
            self.sku_table.setItem(idx, 3, QTableWidgetItem(status))

        self.sku_table.resizeColumnsToContents()

    def _select_worker(self) -> bool:
        """Show worker selection dialog

        Returns:
            bool: True if worker selected, False if cancelled
        """
        try:
            # Show selection dialog
            dialog = WorkerSelectionDialog(self.worker_manager, self)

            if dialog.exec() == QDialog.Accepted:
                self.current_worker_id = dialog.get_selected_worker_id()

                # Get worker details
                worker = self.worker_manager.get_worker(self.current_worker_id)
                if worker:
                    self.current_worker_name = worker.name
                    if hasattr(self, 'sb_worker_label'):
                        self.sb_worker_label.setText(f"Worker: {self.current_worker_name}")
                    logger.info(f"Logged in as: {self.current_worker_name} ({self.current_worker_id})")
                    return True

            logger.info("Worker selection cancelled")
            return False

        except Exception as e:
            logger.error(f"Worker selection failed: {e}", exc_info=True)
            QMessageBox.critical(
                self,
                "Error",
                f"Failed to load worker profiles:\n{str(e)}\n\nApplication will exit."
            )
            return False

    # ========================================================================
    # CLIENT MANAGEMENT (NEW)
    # ========================================================================

    def load_available_clients(self):
        """Load available client profiles and populate dropdown."""
        logger.info("Loading available clients")

        self.client_combo.blockSignals(True)  # Prevent triggering on_client_changed during load
        self.client_combo.clear()

        try:
            clients = self.profile_manager.get_available_clients()

            if not clients:
                logger.warning("No clients found")
                self.client_combo.addItem("(No clients available)", None)
                self.client_combo.setEnabled(False)
                self.status_label.setText("No clients found. Click '+ New Client' to create one.")
                return

            self.client_combo.setEnabled(True)

            for client_id in clients:
                config = self.profile_manager.load_client_config(client_id)
                if config:
                    display_name = f"{config.get('client_name', client_id)} ({client_id})"
                else:
                    display_name = client_id

                self.client_combo.addItem(display_name, client_id)
                logger.debug(f"Added client: {client_id}")

            # Restore last selected client
            last_client = self.settings.value("last_client")
            if last_client:
                index = self.client_combo.findData(last_client)
                if index >= 0:
                    self.client_combo.setCurrentIndex(index)
                    logger.info(f"Restored last selected client: {last_client}")

            logger.info(f"Loaded {len(clients)} clients")

        except Exception as e:
            logger.error(f"Error loading clients: {e}", exc_info=True)
            QMessageBox.warning(self, "Error", f"Failed to load clients:\n\n{e}")

        finally:
            self.client_combo.blockSignals(False)

        # Trigger selection if there's a valid item
        if self.client_combo.currentData():
            self.on_client_changed(self.client_combo.currentIndex())

    def on_client_changed(self, index: int):
        """
        Handle client selection change.

        Args:
            index: Index of selected item in combo box
        """
        client_id = self.client_combo.currentData()

        if not client_id:
            logger.debug("No valid client selected")
            self.current_client_id = None
            self.status_label.setText("Please select or create a client.")
            return

        logger.info(f"Client changed to: {client_id}")

        self.current_client_id = client_id

        # Save as last selected client
        self.settings.setValue("last_client", client_id)

        # Update status
        client_name = self.client_combo.currentText()
        self.status_label.setText(f"Selected client: {client_name}\nReady to start a session.")

        logger.debug(f"Current client set to: {client_id}")

    def create_new_client(self):
        """Open dialog to create a new client profile."""
        logger.info("Opening new client dialog")

        dialog = NewClientDialog(self.profile_manager, self)

        if dialog.exec() == QDialog.Accepted:
            client_id = dialog.client_id
            logger.info(f"New client created: {client_id}")

            # Reload clients and select the new one
            self.load_available_clients()

            # Select newly created client
            index = self.client_combo.findData(client_id)
            if index >= 0:
                self.client_combo.setCurrentIndex(index)

            QMessageBox.information(
                self,
                "Client Created",
                f"Client '{dialog.client_name}' (ID: {client_id}) created successfully!"
            )

    # Muted theme-aware flash colors
    _FLASH_COLORS = {
        "green": "#43a047",
        "red": "#c0392b",
        "orange": "#b06020",
    }
    _FRAME_DEFAULT_STYLE = PackerModeWidget.FRAME_DEFAULT_STYLE

    def flash_border(self, color: str, duration_ms: int = 500):
        """
        Flashes the border of the packer mode table's frame.

        Provides visual feedback for scan results (green = success, red = error).

        Args:
            color (str): Key color: "green", "red", or "orange".
            duration_ms (int): Duration of the flash in milliseconds.
        """
        hex_color = self._FLASH_COLORS.get(color, color)
        self.packer_mode_widget.table_frame.setStyleSheet(
            f"QFrame#TableFrame {{ border: 2px solid {hex_color}; border-radius: 3px; }}"
        )
        QTimer.singleShot(
            duration_ms,
            lambda: self.packer_mode_widget.table_frame.setStyleSheet(self._FRAME_DEFAULT_STYLE)
        )


    def start_session(self, file_path: str = None, restore_dir: str = None):
        """
        Start a new packing session for the currently selected client.

        This method is used to start or restore Shopify Tool sessions. All sessions
        must be created through Shopify Tool - direct Excel file loading is no longer supported.

        Args:
            file_path: Path to the Shopify session directory (contains analysis_data.json).
                      Must be provided - no file dialog shown. Use SessionSelectorDialog
                      to let users choose a session.
            restore_dir: Optional directory of the session to restore (for crash recovery)
        """
        logger.info("Starting new session")

        # Check if client is selected
        if not self.current_client_id:
            logger.warning("Attempted to start session without selecting client")
            self.client_combo.setStyleSheet("border: 2px solid red;")
            QMessageBox.warning(
                self,
                "No Client Selected",
                "Please select a client before starting a session!"
            )
            QTimer.singleShot(2000, lambda: self.client_combo.setStyleSheet(""))
            return

        # Check if session already active
        if self.session_manager and self.session_manager.is_active():
            logger.warning("Attempted to start session while one is already active")
            self.status_label.setText("A session is already active. Please end it first.")
            return

        # Require file_path for Shopify sessions
        if not file_path and not restore_dir:
            logger.error("start_session() called without file_path or restore_dir")
            QMessageBox.warning(
                self,
                "No Session Selected",
                "Please use 'Load Shopify Session' to select a session.\n\n"
                "All sessions must be created through Shopify Tool."
            )
            return

        # Clear existing tree
        if hasattr(self, 'order_tree'):
            self.order_tree.clear()

        logger.info(f"Starting session for client {self.current_client_id} with path: {file_path}")

        try:
            # Create SessionManager for this client
            self.session_manager = SessionManager(
                client_id=self.current_client_id,
                profile_manager=self.profile_manager,
                lock_manager=self.lock_manager,
                worker_id=self.current_worker_id,
                worker_name=self.current_worker_name
            )

            # Start session
            session_id = self.session_manager.start_session(file_path, restore_dir=restore_dir)
            logger.info(f"Session started: {session_id}")

            # Get barcode directory (for Excel workflow backward compatibility)
            # This will be detected as legacy workflow in PackerLogic
            barcodes_dir = self.session_manager.get_barcodes_dir()

            # Create PackerLogic instance
            self.logic = PackerLogic(
                client_id=self.current_client_id,
                profile_manager=self.profile_manager,
                work_dir=barcodes_dir
            )

            # Connect signals
            self.logic.item_packed.connect(self._on_item_packed)
            self.logic.all_orders_complete.connect(self._on_all_orders_complete)

            # Load Shopify session data
            session_path = self.session_manager.output_dir
            order_count, analysis_timestamp = self.logic.load_from_shopify_analysis(session_path)

            logger.info(f"Loaded {order_count} orders from Shopify analysis (analyzed at: {analysis_timestamp})")

            # Setup order table
            self.setup_order_table()

            # Update UI
            self.status_label.setText(f"Successfully loaded {order_count} orders for session '{session_id}'.")
            self.packer_mode_button.setEnabled(True)

        except StaleLockError as e:
            # Session has a stale lock - offer to force-release it
            logger.warning(f"Session has stale lock: {e}")
            self._handle_stale_lock_error(e, file_path, restore_dir)
            self.session_manager = None
            self.logic = None

        except SessionLockedError as e:
            # Session is actively locked by another process
            logger.warning(f"Session is locked: {e}")
            self._handle_session_locked_error(e)
            self.session_manager = None
            self.logic = None

        except Exception as e:
            logger.error(f"Failed to start session: {e}", exc_info=True)
            QMessageBox.critical(self, "Error", f"Failed to start session:\n\n{e}")
            if self.session_manager:
                self.session_manager.end_session()
            self.session_manager = None
            self.logic = None

    def _toggle_theme(self):
        """Toggle between dark and light themes."""
        from PySide6.QtWidgets import QApplication
        new_theme = toggle_theme(QApplication.instance())
        self.statusBar().showMessage(f"Theme switched to: {new_theme}", 3000)

    def open_sku_mapping_dialog(self):
        """
        Open SKU mapping dialog for current client.

        Phase 1.3: Uses ProfileManager for centralized storage on file server.
        All changes are synchronized across all PCs with file locking.
        """
        if not self.current_client_id:
            logger.warning("Attempted to open SKU mapping without selecting client")
            QMessageBox.warning(
                self,
                "No Client Selected",
                "Please select a client first!"
            )
            return

        logger.info(f"Opening SKU mapping dialog for client {self.current_client_id}")

        # Phase 1.3: Use ProfileManager directly for centralized storage
        dialog = SKUMappingDialog(self.current_client_id, self.profile_manager, self)

        if dialog.exec():  # User clicked "Save & Close"
            # Mappings are already saved by the dialog
            logger.info("SKU mapping dialog closed with save")

            # If a session is active, reload the SKU map into logic instance
            if self.logic:
                try:
                    new_map = self.profile_manager.load_sku_mapping(self.current_client_id)
                    self.logic.set_sku_map(new_map)
                    self.status_label.setText("SKU mapping updated and synchronized across all PCs.")
                    logger.info("SKU mapping reloaded into active session")
                except Exception as e:
                    logger.error(f"Failed to reload SKU mapping into session: {e}")
                    QMessageBox.warning(
                        self,
                        "Reload Warning",
                        f"Mappings saved successfully but failed to reload into current session:\n\n{e}\n\n"
                        f"Please restart the session to use new mappings."
                    )

    def _start_heartbeat_timer(self):
        """Start timer to update session lock heartbeat."""
        if hasattr(self, 'heartbeat_timer'):
            self.heartbeat_timer.stop()

        self.heartbeat_timer = QTimer(self)
        self.heartbeat_timer.timeout.connect(self._update_session_heartbeat)
        self.heartbeat_timer.start(60000)  # 60 seconds
        logger.debug("Heartbeat timer started")

    def _update_session_heartbeat(self):
        """Update heartbeat for active session lock."""
        if self.logic and hasattr(self, 'current_work_dir') and self.current_work_dir:
            try:
                self.lock_manager.update_heartbeat(Path(self.current_work_dir))
                logger.debug("Lock heartbeat updated")
            except Exception as e:
                logger.error(f"Failed to update heartbeat: {e}")

    def _cleanup_failed_session_start(self):
        """
        Clean up resources after failed session start.
        Extracted to avoid code duplication in exception handlers.
        """
        # Stop heartbeat timer if running
        if hasattr(self, 'heartbeat_timer') and self.heartbeat_timer:
            try:
                self.heartbeat_timer.stop()
                logger.debug("Heartbeat timer stopped in cleanup")
            except Exception as timer_error:
                logger.warning(f"Failed to stop heartbeat timer: {timer_error}")

        # Release lock if acquired
        if hasattr(self, 'current_work_dir') and self.current_work_dir:
            try:
                self.lock_manager.release_lock(Path(self.current_work_dir))
                logger.info(f"Lock released during cleanup: {self.current_work_dir}")
            except Exception as lock_error:
                logger.warning(f"Failed to release lock: {lock_error}")

        # Clear state
        if hasattr(self, 'logic') and self.logic:
            self.logic = None

        # Clear instance variables
        if hasattr(self, 'current_work_dir'):
            self.current_work_dir = None
        if hasattr(self, 'current_session_path'):
            self.current_session_path = None
        if hasattr(self, 'current_packing_list'):
            self.current_packing_list = None
        if hasattr(self, 'packing_data'):
            self.packing_data = None

    def closeEvent(self, event: QCloseEvent):
        """
        Handle application close event - ensure clean shutdown.

        This method is called when:
        - User clicks X button
        - User presses Alt+F4
        - Application receives SIGTERM
        - System shutdown initiated

        Critical for multi-PC warehouse deployment:
        - Releases session locks immediately
        - Stops heartbeat timers
        - Saves session state
        - Prevents 2-minute stale lock timeout

        Args:
            event: Qt close event
        """
        logger.info("Application closing, performing cleanup...")

        try:
            # 1. Stop heartbeat timer (prevents lock updates during cleanup)
            if hasattr(self, 'heartbeat_timer') and self.heartbeat_timer:
                try:
                    self.heartbeat_timer.stop()
                    logger.info("Heartbeat timer stopped")
                except Exception as e:
                    logger.warning(f"Failed to stop heartbeat timer: {e}")

            # 2. Save current packing state (if session active)
            if hasattr(self, 'logic') and self.logic:
                try:
                    self.logic.save_state()
                    logger.info("Packing state saved")
                except Exception as e:
                    logger.warning(f"Failed to save packing state: {e}")

            # 3. Release lock on current work directory
            if hasattr(self, 'current_work_dir') and self.current_work_dir:
                try:
                    self.lock_manager.release_lock(Path(self.current_work_dir))
                    logger.info(f"Lock released: {self.current_work_dir}")
                except Exception as e:
                    logger.warning(f"Failed to release lock: {e}")

            # 4. End Excel session if active
            if hasattr(self, 'session_manager') and self.session_manager:
                try:
                    if self.session_manager.is_active():
                        # Close without generating full report (unexpected close)
                        # The session can be resumed later via Session Browser
                        logger.info("Excel session detected, closing gracefully")
                except Exception as e:
                    logger.warning(f"Failed to check session manager: {e}")

            # 5. Stop auto-refresh timer in Session Browser if open
            if hasattr(self, 'session_browser_dialog'):
                try:
                    if hasattr(self.session_browser_dialog, 'auto_refresh_timer'):
                        self.session_browser_dialog.auto_refresh_timer.stop()
                        logger.info("Session browser auto-refresh stopped")
                except Exception as e:
                    logger.warning(f"Failed to stop session browser timer: {e}")

            logger.info("Application cleanup completed successfully")

        except Exception as e:
            # Log but don't prevent shutdown
            logger.error(f"Error during application cleanup: {e}", exc_info=True)

        # Always accept the event (allow application to close)
        event.accept()

    def start_shopify_packing_session(
        self,
        packing_list_path: Path,
        work_dir: Path,
        session_path: Path,
        client_id: str,
        packing_list_name: str
    ) -> bool:
        """
        Start packing session for Shopify packing list (new or resumed).

        This method handles BOTH:
        - Resuming interrupted sessions (from Session Browser)
        - Starting new sessions (from open_shopify_session)

        Args:
            packing_list_path: Path to packing list JSON file
            work_dir: Path to work directory (packing/{list_name}/)
            session_path: Path to session directory (Sessions/CLIENT_X/2025-XX-XX_X/)
            client_id: Client identifier
            packing_list_name: Name of the packing list

        Returns:
            bool: True if session started successfully, False otherwise

        Raises:
            FileNotFoundError: If packing list file not found
            json.JSONDecodeError: If packing list JSON is invalid
            ValueError: If packing data is invalid
            RuntimeError: If barcode generation fails
        """
        try:
            logger.info(f"Starting Shopify packing session: {packing_list_path}")
            logger.info(f"Work directory: {work_dir}")
            logger.info(f"Session path: {session_path}")

            # 1. Validate packing list file exists
            if not packing_list_path.exists():
                raise FileNotFoundError(f"Packing list not found: {packing_list_path}")

            # 2. Store session state
            self.current_session_path = str(session_path)
            self.current_packing_list = packing_list_name
            self.current_work_dir = str(work_dir)

            # 3. Acquire lock on work directory (with stale lock handling)
            success, error_msg, _ = self.lock_manager.acquire_lock(
                client_id=client_id,
                session_dir=work_dir,
                worker_id=self.current_worker_id,
                worker_name=self.current_worker_name
            )

            if not success:
                # Check if stale lock (error message contains "stale" keyword)
                if error_msg and "stale" in error_msg.lower():
                    reply = QMessageBox.question(
                        self, "Stale Lock Detected",
                        f"{error_msg}\n\nForce-release lock and continue?",
                        QMessageBox.Yes | QMessageBox.No
                    )
                    if reply == QMessageBox.Yes:
                        self.lock_manager.force_release_lock(work_dir)
                        success, error_msg, _ = self.lock_manager.acquire_lock(
                            client_id,
                            work_dir,
                            worker_id=self.current_worker_id,
                            worker_name=self.current_worker_name
                        )
                        if not success:
                            raise RuntimeError(f"Failed to acquire lock after force-release: {error_msg}")
                    else:
                        # User chose not to force-release
                        return False
                else:
                    # Active lock by another user
                    raise RuntimeError(error_msg or "Session is locked by another user")

            logger.info(f"Lock acquired on {work_dir}")

            # 4. Start heartbeat timer
            self._start_heartbeat_timer()
            logger.info("Heartbeat timer started")

            # 5 & 7. Initialize PackerLogic + load packing list in background thread
            # so the UI remains responsive (progress dialog animates while server is slow).
            progress = QProgressDialog("Loading packing list…", None, 0, 0, self)
            progress.setWindowTitle("Please Wait")
            progress.setWindowModality(Qt.WindowModal)
            progress.setCancelButton(None)
            progress.setMinimumDuration(0)
            progress.setValue(0)
            progress.show()
            QApplication.processEvents()

            start_worker = SessionStartWorker(
                client_id=client_id,
                profile_manager=self.profile_manager,
                work_dir=work_dir,
                packing_list_path=packing_list_path,
                parent=self,
            )
            start_worker.start()
            while not start_worker.wait(50):
                QApplication.processEvents()
            progress.close()

            if start_worker.error is not None:
                raise start_worker.error

            self.logic = start_worker.logic
            order_count = start_worker.order_count
            list_name = start_worker.list_name

            # 6. Connect signals (must happen on main thread after moveToThread)
            self.logic.item_packed.connect(self._on_item_packed)
            self.logic.all_orders_complete.connect(self._on_all_orders_complete)

            logger.info(f"Loaded {order_count} orders from packing list")

            # 8. Set minimal packing data for UI
            self.packing_data = {
                'list_name': list_name,
                'total_orders': order_count,
                'orders': []  # Don't duplicate data - PackerLogic has it
            }

            # 9. Update session metadata
            if hasattr(self.session_manager, 'update_session_metadata'):
                try:
                    self.session_manager.update_session_metadata(
                        self.current_session_path,
                        self.current_packing_list,
                        'in_progress'
                    )
                except Exception as e:
                    logger.warning(f"Could not update session metadata: {e}")

            # 9b. Register session start in registry (enables fast browser loading)
            try:
                _reg_total_items = 0
                if self.logic and self.logic.processed_df is not None:
                    _reg_total_items = int(
                        pd.to_numeric(
                            self.logic.processed_df['Quantity'], errors='coerce'
                        ).fillna(0).sum()
                    )
                self.registry_manager.register_session_start(
                    client_id=client_id,
                    session_id=session_path.name,
                    packing_list_name=packing_list_name,
                    worker_id=self.current_worker_id,
                    worker_name=self.current_worker_name,
                    pc_name=os.environ.get('COMPUTERNAME', 'Unknown'),
                    total_orders=order_count,
                    total_items=_reg_total_items,
                    work_dir=str(work_dir),
                    session_path=str(session_path),
                )
            except Exception as _e:
                logger.warning(f"Registry update (session start) failed: {_e}")

            # 10. Setup order table
            self.setup_order_table()

            # 11. Update UI state
            self.status_label.setText(
                f"Session: {session_path.name} / {packing_list_name}\n"
                f"Orders: {order_count}\n"
                f"Ready for packing"
            )

            # 12. Enable packing UI
            self.enable_packing_mode()

            logger.info("Shopify packing session started successfully")
            return True

        except FileNotFoundError as e:
            logger.error(f"Packing list file not found: {e}", exc_info=True)
            self._cleanup_failed_session_start()
            QMessageBox.critical(
                self,
                "File Not Found",
                f"Packing list file not found:\n{str(e)}"
            )
            return False

        except json.JSONDecodeError as e:
            logger.error(f"Invalid JSON in packing list: {e}", exc_info=True)
            self._cleanup_failed_session_start()
            QMessageBox.critical(
                self,
                "Invalid JSON",
                f"Packing list contains invalid JSON:\n{str(e)}"
            )
            return False

        except ValueError as e:
            logger.error(f"Invalid packing data: {e}", exc_info=True)
            self._cleanup_failed_session_start()
            QMessageBox.critical(
                self,
                "Invalid Data",
                f"Packing list contains invalid data:\n{str(e)}"
            )
            return False

        except RuntimeError as e:
            logger.error(f"Failed to start session: {e}", exc_info=True)
            self._cleanup_failed_session_start()
            QMessageBox.critical(
                self,
                "Session Start Failed",
                f"Failed to start packing session:\n{str(e)}"
            )
            return False

        except Exception as e:
            logger.error(f"Unexpected error starting session: {e}", exc_info=True)
            self._cleanup_failed_session_start()
            QMessageBox.critical(
                self,
                "Error",
                f"Unexpected error starting packing session:\n{str(e)}"
            )
            return False

    def end_session(self):
        """
        Ends the current session gracefully.

        This involves saving a final report, cleaning up session files, and
        resetting the UI to its initial state.

        For Shopify sessions (unified work directory):
        - Report saved to: current_work_dir/reports/packing_completed.xlsx

        For Excel sessions (legacy):
        - Report saved to: session_dir/[original_filename]_completed.xlsx
        """
        # Check if any session is active (Excel or Shopify)
        is_excel_session = self.session_manager and self.session_manager.is_active()
        is_shopify_session = hasattr(self, 'current_work_dir') and self.current_work_dir

        if not (is_excel_session or is_shopify_session):
            logger.warning("end_session called but no active session found")
            return

        try:
            # Determine output path based on session type
            if hasattr(self, 'current_work_dir') and self.current_work_dir:
                # Shopify session - save to unified work directory
                report_dir = Path(self.current_work_dir) / "reports"
                report_dir.mkdir(exist_ok=True, parents=True)
                new_filename = "packing_completed.xlsx"
                output_path = str(report_dir / new_filename)
                logger.info(f"Saving Shopify session report to: {output_path}")
            else:
                # Excel session - save to session directory (legacy behavior)
                output_dir = self.session_manager.get_output_dir()
                original_filename = os.path.basename(self.session_manager.packing_list_path)
                new_filename = f"{os.path.splitext(original_filename)[0]}_completed.xlsx"
                output_path = os.path.join(output_dir, new_filename)
                logger.info(f"Saving Excel session report to: {output_path}")

            # Generate status map from session packing state
            completed_orders_set = set(self.logic.session_packing_state.get('completed_orders', []))
            in_progress_orders = self.logic.session_packing_state.get('in_progress', {})

            final_df = self.logic.packing_list_df.copy()

            # Add Status column
            final_df['Status'] = final_df['Order_Number'].apply(
                lambda x: 'Completed' if x in completed_orders_set else ('In Progress' if x in in_progress_orders else 'New')
            )

            # Add Completed At column
            final_df['Completed At'] = final_df['Order_Number'].apply(
                lambda x: datetime.now().strftime("%Y-%m-%d %H:%M:%S") if x in completed_orders_set else ''
            )

            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                final_df.to_excel(writer, index=False, sheet_name='Sheet1')
                workbook = writer.book
                worksheet = writer.sheets['Sheet1']
                green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                status_col_idx = final_df.columns.get_loc('Status') + 1

                for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row)):
                    status_cell = worksheet.cell(row=row_idx + 2, column=status_col_idx)
                    if status_cell.value == 'Completed':
                        for cell in row:
                            cell.fill = green_fill

            self.status_label.setText(f"Session ended. Report saved to {output_path}")

            # Generate session summary + record stats in background thread so the
            # UI stays responsive while writing to the (potentially slow) file server.
            if self.logic:
                _is_shopify = hasattr(self, 'current_work_dir') and self.current_work_dir

                if _is_shopify:
                    _summary_path = os.path.join(self.current_work_dir, "session_summary.json")
                    _session_type = "shopify"
                else:
                    _barcodes_dir = self.session_manager.get_barcodes_dir()
                    _summary_path = os.path.join(_barcodes_dir, "session_summary.json")
                    _session_type = "excel"

                # --- Gather all stats data on the main thread (fast, no server I/O) ---
                try:
                    _session_info = self.session_manager.get_session_info()
                    _start_time = None
                    if _session_info and 'started_at' in _session_info:
                        try:
                            _start_time = datetime.fromisoformat(_session_info['started_at'])
                        except (ValueError, TypeError):
                            logger.warning("Could not parse started_at from session_info")
                except Exception as e:
                    logger.warning(f"Could not get session_info: {e}")
                    _session_info = None
                    _start_time = None

                _end_time = datetime.now()
                _completed_orders_list = self.logic.session_packing_state.get('completed_orders', [])
                _completed_orders = len(_completed_orders_list)
                _in_progress_orders_dict = self.logic.session_packing_state.get('in_progress', {})
                _in_progress_orders = len(_in_progress_orders_dict)

                _items_packed = 0
                try:
                    if self.logic.processed_df is not None and _completed_orders_list:
                        _ci = pd.to_numeric(
                            self.logic.processed_df[
                                self.logic.processed_df['Order_Number'].isin(_completed_orders_list)
                            ]['Quantity'],
                            errors='coerce'
                        ).sum()
                        _items_packed += int(_ci)
                    for _osl in _in_progress_orders_dict.values():
                        if isinstance(_osl, list):
                            for _sd in _osl:
                                if isinstance(_sd, dict):
                                    _items_packed += _sd.get('packed', 0)
                except Exception as e:
                    logger.error(f"Error calculating items_packed: {e}", exc_info=True)

                _total_orders, _total_items = 0, 0
                try:
                    if self.logic.processed_df is not None:
                        _total_orders = len(self.logic.processed_df['Order_Number'].unique())
                        _total_items = int(pd.to_numeric(self.logic.processed_df['Quantity'], errors='coerce').sum())
                except Exception as e:
                    logger.error(f"Error calculating totals: {e}", exc_info=True)

                if _is_shopify:
                    _session_id = f"{getattr(self, 'current_session_path', '')}_{getattr(self, 'current_packing_list', '')}"
                    _pl_path_str = getattr(self, 'current_packing_list', 'Unknown') or 'Unknown'
                else:
                    _session_id = self.session_manager.session_id
                    _pl_path_str = str(self.session_manager.packing_list_path or 'Unknown')

                _duration_seconds = int((_end_time - _start_time).total_seconds()) if _start_time else None

                # Capture non-Qt references for the closure
                _logic_ref = self.logic
                _client_id = self.current_client_id
                _worker_id = self.current_worker_id
                _worker_name = self.current_worker_name
                _stats_mgr = self.stats_manager
                _worker_mgr = self.worker_manager
                _sess_mgr = self.session_manager
                _cur_sess_path = getattr(self, 'current_session_path', None)
                _cur_pack_list = getattr(self, 'current_packing_list', None)
                _registry_mgr = getattr(self, 'registry_manager', None)

                # Flush any pending state write on the main thread *before*
                # handing off to the background worker.  AsyncStateWriter's
                # flush() must only be called from the main/UI thread.
                _logic_ref._state_writer.flush()

                def _do_slow_writes():
                    # 1. Save session summary
                    try:
                        _logic_ref.save_session_summary(
                            summary_path=_summary_path,
                            worker_id=_worker_id,
                            worker_name=_worker_name,
                            session_type=_session_type,
                        )
                        logger.info(f"Session summary saved to: {_summary_path}")
                    except Exception as exc:
                        logger.error(f"save_session_summary failed: {exc}", exc_info=True)
                        try:
                            from shared.metadata_utils import get_current_timestamp
                            _minimal = {
                                "version": "1.3.0",
                                "session_id": _logic_ref.session_id if _logic_ref else "unknown",
                                "session_type": _session_type,
                                "client_id": _client_id,
                                "worker_id": _worker_id,
                                "worker_name": _worker_name,
                                "completed_at": get_current_timestamp(),
                                "error": str(exc),
                            }
                            with open(_summary_path, 'w', encoding='utf-8') as _f:
                                import json as _json
                                _json.dump(_minimal, _f, indent=2, ensure_ascii=False)
                        except Exception:
                            pass

                    # 2. Record to stats
                    try:
                        _stats_mgr.record_packing(
                            client_id=_client_id,
                            session_id=_session_id,
                            worker_id=_worker_id,
                            orders_count=_completed_orders,
                            items_count=_items_packed,
                            metadata={
                                "duration_seconds": _duration_seconds,
                                "packing_list_name": os.path.basename(_pl_path_str),
                                "started_at": _start_time.isoformat() if _start_time else None,
                                "completed_at": _end_time.isoformat(),
                                "total_orders": _total_orders,
                                "in_progress_orders": _in_progress_orders,
                                "session_type": _session_type,
                                "user_name": os.environ.get('USERNAME', 'Unknown'),
                                "worker_name": _worker_name,
                                "pc_name": os.environ.get('COMPUTERNAME', 'Unknown'),
                            },
                        )
                        logger.info(f"Recorded {_completed_orders} orders, {_items_packed} items to stats")
                    except Exception as exc:
                        logger.error(f"record_packing failed: {exc}", exc_info=True)

                    # 3. Update worker stats
                    try:
                        if _worker_id:
                            _worker_mgr.update_worker_stats(
                                worker_id=_worker_id,
                                sessions=1,
                                orders=_completed_orders,
                                items=_items_packed,
                                duration_seconds=_duration_seconds or 0,
                                session_id=_session_id,
                            )
                            logger.info(f"Updated worker stats for {_worker_name}")
                    except Exception as exc:
                        logger.error(f"update_worker_stats failed: {exc}", exc_info=True)

                    # 4. Update session metadata
                    try:
                        if _cur_sess_path and _cur_pack_list:
                            _sess_mgr.update_session_metadata(
                                _cur_sess_path, _cur_pack_list, 'completed'
                            )
                            logger.info("Updated session metadata to 'completed'")
                    except Exception as exc:
                        logger.warning(f"update_session_metadata failed: {exc}")

                    # 5. Update session registry with completed status + metrics
                    try:
                        if _registry_mgr and _client_id and _cur_sess_path and _cur_pack_list:
                            import json as _json_reg
                            with open(_summary_path, 'r', encoding='utf-8') as _f_reg:
                                _reg_summary = _json_reg.load(_f_reg)
                            _session_id_reg = Path(_cur_sess_path).name
                            _registry_mgr.register_session_complete(
                                _client_id,
                                _session_id_reg,
                                _cur_pack_list,
                                _reg_summary,
                            )
                            logger.info("Registry updated with session completion")
                    except Exception as exc:
                        logger.warning(f"Registry update (session complete) failed: {exc}")

                # Show progress dialog while writes happen in background
                _end_progress = QProgressDialog("Saving session…", None, 0, 0, self)
                _end_progress.setWindowTitle("Please Wait")
                _end_progress.setWindowModality(Qt.WindowModal)
                _end_progress.setCancelButton(None)
                _end_progress.setMinimumDuration(0)
                _end_progress.setValue(0)
                _end_progress.show()
                QApplication.processEvents()

                _end_worker = SessionEndWorker(_do_slow_writes, self)
                _end_worker.start()
                while not _end_worker.wait(50):
                    QApplication.processEvents()
                _end_progress.close()

                if _end_worker.error:
                    logger.error(f"Session end writes had an error: {_end_worker.error}")

        except Exception as e:
            self.status_label.setText(f"Could not save the report. Error: {e}")
            logger.error(f"Error during end_session: {e}")

        # ✅ CRITICAL: Stop heartbeat timer and release lock
        if hasattr(self, 'heartbeat_timer'):
            self.heartbeat_timer.stop()
            logger.debug("Heartbeat timer stopped")

        if hasattr(self, 'current_work_dir') and self.current_work_dir:
            try:
                self.lock_manager.release_lock(Path(self.current_work_dir))
                logger.info("Lock released")
            except Exception as e:
                logger.error(f"Failed to release lock: {e}")

        # Cleanup PackerLogic state
        if self.logic:
            self.logic.end_session_cleanup()
            self.logic = None

        # End Excel session if active
        if self.session_manager and self.session_manager.is_active():
            self.session_manager.end_session()

        # Clear Shopify session variables
        if hasattr(self, 'current_work_dir'):
            self.current_work_dir = None
        if hasattr(self, 'current_session_path'):
            self.current_session_path = None
        if hasattr(self, 'current_packing_list'):
            self.current_packing_list = None
        if hasattr(self, 'packing_data'):
            self.packing_data = None

        self.packer_mode_button.setEnabled(False)

        # Disable toolbar end button and reset session info
        if hasattr(self, 'toolbar_end_btn'):
            self.toolbar_end_btn.setEnabled(False)
        if hasattr(self, 'session_info_label'):
            self.session_info_label.setText("No active session")

        if hasattr(self, 'order_tree'):
            self.order_tree.clear()
        self.status_label.setText("Session ended. Start a new session to begin.")

        # Return user to session view (avoids leaving a blank packer mode screen)
        if hasattr(self, 'stacked_widget') and hasattr(self, 'session_widget'):
            self.stacked_widget.setCurrentWidget(self.session_widget)

        logger.info("Session ended and all variables cleared")

    def view_session_history(self):
        """
        View completed packing sessions history.

        TODO: Implement UI for viewing historical session data:
        - List all completed sessions from Sessions/CLIENT_*/
        - Show session_summary.json data (orders completed, duration, performance)
        - Filter by date range, client, packing list name
        - Display statistics: average orders/hour, total sessions, etc.
        - Link to packing_state.json for detailed review

        This feature will use:
        - session_summary.json: Aggregated statistics and performance metrics
        - packing_state.json: Detailed per-order packing history
        - session_info.json: Session metadata (PC name, start time, etc.)

        Planned UI:
        - Table view with sortable columns
        - Date range picker
        - Export to CSV/Excel
        - Charts for performance trends
        """
        logger.info("view_session_history called (placeholder - not yet implemented)")

        QMessageBox.information(
            self,
            "Feature Coming Soon",
            "Session History viewer is coming soon!\n\n"
            "This feature will show:\n"
            "• Completed packing sessions\n"
            "• Performance metrics (orders/hour, items/hour)\n"
            "• Historical trends and statistics\n"
            "• Detailed session data\n\n"
            "Stay tuned!"
        )


    def setup_order_table(self):
        """
        Sets up the expandable order tree and statistics display.

        This method populates the tree widget with orders and items,
        and updates the statistics tab with session metrics.
        """
        # Populate the expandable tree
        self._populate_order_tree()

        # Update statistics
        self._update_statistics()

        logger.info("Order tree and statistics updated successfully")

    def switch_to_packer_mode(self):
        """Switches the view to the Packer Mode widget."""
        self.stacked_widget.setCurrentWidget(self.packer_mode_widget)
        self.packer_mode_widget.set_focus_to_scanner()

    def switch_to_session_view(self):
        """Switches the view back to the main session widget (tabbed interface)."""
        if self.logic:
            self.logic.clear_current_order()
        if self.packer_mode_widget:
            self.packer_mode_widget.clear_screen()
        self.stacked_widget.setCurrentWidget(self.session_widget)

    def open_print_dialog(self):
        """Opens the dialog for printing order barcodes."""
        if not self.logic.orders_data:
            self.status_label.setText("No data to print.")
            return
        dialog = PrintDialog(self.logic.orders_data, self)
        dialog.exec()

    def on_scanner_input(self, text: str):
        """
        Handles input from the barcode scanner in Packer Mode.

        This is the central callback for all barcode scans. It determines if
        the scan is for an order or a product SKU and routes the logic accordingly.

        Args:
            text (str): The decoded text from the barcode scanner.
        """
        self.packer_mode_widget.update_raw_scan_display(text)
        self.packer_mode_widget.show_notification("", "transparent")

        if self.logic.current_order_number is None:
            # Find order by normalized comparison
            order_number_from_scan = None
            scanned_normalized = self.logic._normalize_order_number(text)
            for order_num in self.logic.orders_data.keys():
                if self.logic._normalize_order_number(order_num) == scanned_normalized:
                    order_number_from_scan = order_num
                    break

            if not order_number_from_scan:
                self.packer_mode_widget.show_notification("ORDER NOT FOUND", "#c0392b")
                self.flash_border("red")
                return

            if order_number_from_scan in self.logic.session_packing_state.get('completed_orders', []):
                self.packer_mode_widget.show_notification(f"ORDER {order_number_from_scan} ALREADY COMPLETED", "#b06020")
                self.flash_border("orange")
                return

            items, status = self.logic.start_order_packing(text)
            if status == "ORDER_LOADED":
                self.packer_mode_widget.add_order_to_history(order_number_from_scan)
                order_metadata = self.logic.orders_data.get(
                    order_number_from_scan, {}
                ).get('metadata', {})
                self.packer_mode_widget.display_order(
                    items,
                    self.logic.current_order_state,
                    metadata=order_metadata,
                    sku_map=self.logic.sku_map,
                )
                completed = len(self.logic.session_packing_state.get('completed_orders', []))
                self.packer_mode_widget.update_session_progress(completed, len(self.logic.orders_data))
                self.update_order_status(order_number_from_scan, "In Progress")
                _beep(1000, 120)
            else:
                self.packer_mode_widget.show_notification("ORDER NOT FOUND", "#c0392b")
                self.flash_border("red")
                _beep(400, 350)
        else:
            result, status = self.logic.process_sku_scan(text)
            if status == "SKU_OK":
                self.packer_mode_widget.update_item_row(result["row"], result["packed"], result["is_complete"])
                self.packer_mode_widget.show_notification("ITEM OK", "#43a047")
                self.flash_border("green")
                _beep(1200, 80)
            elif status == "SKU_NOT_FOUND":
                unknown_list = self.logic.unknown_scans
                if len(unknown_list) > 1:
                    detail = f"({len(unknown_list)} unknown scans)\nLast: {text}"
                else:
                    detail = f"Unknown: {text}"
                self.packer_mode_widget.show_notification(
                    f"INCORRECT ITEM!\n{detail}", "#c0392b"
                )
                self.flash_border("red")
                _beep(400, 350)
            elif status == "SKU_EXTRA":
                self.packer_mode_widget.show_notification("EXTRA ITEM!", "#b06020")
                self.flash_border("orange")
                _beep(700, 200)
                self.packer_mode_widget.show_extras_panel(self.logic.current_extra_items)
            elif status == "ORDER_COMPLETE_WITH_EXTRAS":
                self.packer_mode_widget.update_item_row(result["row"], result["packed"], result["is_complete"])
                self.packer_mode_widget.show_notification("REVIEW EXTRA ITEMS!", "#e67e22")
                self.flash_border("orange")
                self.packer_mode_widget.show_extras_panel(self.logic.current_extra_items)
            elif status == "ORDER_COMPLETE":
                current_order_num = self.logic.current_order_number
                self.packer_mode_widget.update_item_row(result["row"], result["packed"], result["is_complete"])
                self._handle_order_completion(current_order_num)
                self.logic.clear_current_order()

    # REMOVED: _process_shopify_packing_data() method (dead code)
    # This method was never called. Functionality replaced by PackerLogic.load_packing_list_json()
    # which is used in start_shopify_packing_session()

    def _on_item_packed(self, order_number: str, packed_count: int, required_count: int):
        """
        Slot to handle real-time progress updates from the logic layer.

        This method is connected to the `item_packed` signal from PackerLogic.
        It refreshes the tree and statistics to reflect the updated progress.

        Args:
            order_number (str): The order number that was updated.
            packed_count (int): The new total of items packed for the order.
            required_count (int): The total items required for the order.
        """
        # Refresh tree and statistics to show updated progress
        self._populate_order_tree()
        self._update_statistics()
        logger.debug(f"Order {order_number} progress: {packed_count}/{required_count}")

    def update_order_status(self, order_number: str, status: str):
        """
        Updates the status of an order in the tree view.

        Args:
            order_number (str): The order number to update.
            status (str): The new status ('In Progress' or 'Completed').
        """
        # Simply refresh the tree and statistics to reflect the new status
        self._populate_order_tree()
        self._update_statistics()
        logger.debug(f"Order {order_number} status updated to: {status}")

    # ─── Packer Mode new action handlers ─────────────────────────────────────

    def _handle_order_completion(self, order_number: str):
        """Shared teardown for every order-complete path (scan, force confirm, extra resolve)."""
        self.packer_mode_widget.show_notification(f"ORDER {order_number} COMPLETE!", "#43a047")
        self.flash_border("green")
        _beep(1200, 80)
        QTimer.singleShot(180, lambda: _beep(1200, 80))
        self.update_order_status(order_number, "Completed")
        if self.logic:
            completed = len(self.logic.session_packing_state.get('completed_orders', []))
            self.packer_mode_widget.update_session_progress(completed, len(self.logic.orders_data))
        self.packer_mode_widget.scanner_input.setEnabled(False)
        QTimer.singleShot(3000, self.packer_mode_widget.clear_screen)

    def _on_skip_order(self):
        """Skip the currently active order (preserves packing progress for later)."""
        if not self.logic or not self.logic.current_order_number:
            return
        skipped = self.logic.current_order_number
        self.logic.skip_order()
        self.packer_mode_widget.add_order_to_history(skipped, "[SKIPPED]")
        self.packer_mode_widget.clear_screen()
        logger.info(f"Order {skipped} skipped")

    def _on_cancel_item(self, row: int):
        """Handle -1 (undo last scan) for a specific item row."""
        if not self.logic:
            return
        result, status = self.logic.cancel_item_scan(row)
        if status == "ITEM_DECREMENTED":
            self.packer_mode_widget.update_item_row(row, result["packed"], False)
            self.flash_border("orange")
        elif status == "ITEM_ALREADY_ZERO":
            self.packer_mode_widget.show_notification("Already at 0!", "#b06020")
        self.packer_mode_widget.set_focus_to_scanner()

    def _on_force_confirm(self, row: int):
        """Handle Force Confirm for a specific item row."""
        if not self.logic:
            return
        result, status = self.logic.force_confirm_item(row)
        if status == "FORCE_CONFIRMED":
            self.packer_mode_widget.update_item_row(row, result["packed"], True)
            if result.get("order_complete"):
                self.flash_border("green")
                order_num = self.logic.current_order_number
                self._handle_order_completion(order_num)
                self.logic.clear_current_order()
            elif self.logic.current_extra_items:
                # All items packed but extra items need resolution before completing
                self.flash_border("orange")
                self.packer_mode_widget.show_notification("REVIEW EXTRA ITEMS!", "#e67e22")
                self.packer_mode_widget.show_extras_panel(self.logic.current_extra_items)
            else:
                self.flash_border("green")
        self.packer_mode_widget.set_focus_to_scanner()

    def _on_map_sku_from_packer(self, sku: str):
        """Quick-add barcode→SKU mapping from packer mode.

        Pre-populates the SKU so the worker only needs to scan/type the barcode.
        The barcode field is left empty for the scanner to fill in.
        """
        if not self.current_client_id:
            self.packer_mode_widget.set_focus_to_scanner()
            return

        barcode, ok = QInputDialog.getText(
            self,
            "Map Barcode to SKU",
            f"SKU:  {sku}\n\nScan or type the product barcode to map to this SKU:",
        )
        if not (ok and barcode and barcode.strip()):
            self.packer_mode_widget.set_focus_to_scanner()
            return

        barcode = barcode.strip()
        try:
            existing = self.profile_manager.load_sku_mapping(self.current_client_id)
            if barcode in existing and existing[barcode] != sku:
                reply = QMessageBox.question(
                    self,
                    "Overwrite Mapping?",
                    f"Barcode '{barcode}' already maps to '{existing[barcode]}'.\n\n"
                    f"Replace with '{sku}'?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                    QMessageBox.StandardButton.No,
                )
                if reply != QMessageBox.StandardButton.Yes:
                    self.packer_mode_widget.set_focus_to_scanner()
                    return

            existing[barcode] = sku
            success = self.profile_manager.save_sku_mapping(self.current_client_id, existing)
            if success:
                if self.logic:
                    self.logic.sku_map = {
                        self.logic._normalize_sku(k): v for k, v in existing.items()
                    }
                    logger.info(f"Quick-mapped barcode '{barcode}' → SKU '{sku}'")
                self.packer_mode_widget.show_notification(f"Mapped: {barcode} → {sku}", "#43a047")
            else:
                QMessageBox.warning(self, "Save Failed", "Could not save mapping to file server.")
        except Exception as e:
            logger.error(f"Failed to save quick SKU mapping: {e}")
            QMessageBox.critical(self, "Error", f"Failed to save mapping:\n\n{e}")

        self.packer_mode_widget.set_focus_to_scanner()

    def _on_extra_confirmed(self, norm_sku: str):
        """Handle 'Keep' for an extra item — user acknowledges it is intentional."""
        if not self.logic:
            return
        _, status = self.logic.confirm_keep_extra(norm_sku)
        self.packer_mode_widget.show_extras_panel(self.logic.current_extra_items)
        if status == "ORDER_NOW_COMPLETE":
            order_num = self.logic.current_order_number
            self._handle_order_completion(order_num)
            self.logic.clear_current_order()
        elif status == "EXTRA_CLEARED":
            self.packer_mode_widget.show_notification(
                "Extra cleared — continue scanning", "#43a047"
            )
        self.packer_mode_widget.set_focus_to_scanner()

    def _on_extra_removed(self, norm_sku: str):
        """Handle 'Remove' for an extra item — user acknowledges it was a mistake."""
        if not self.logic:
            return
        _, status = self.logic.remove_extra_item(norm_sku)
        self.packer_mode_widget.show_extras_panel(self.logic.current_extra_items)
        if status == "ORDER_NOW_COMPLETE":
            order_num = self.logic.current_order_number
            self._handle_order_completion(order_num)
            self.logic.clear_current_order()
        elif status == "EXTRA_CLEARED":
            self.packer_mode_widget.show_notification(
                "Extra cleared — continue scanning", "#43a047"
            )
        self.packer_mode_widget.set_focus_to_scanner()

    def _on_all_orders_complete(self):
        """Handle the all_orders_complete signal — defer dialog to next event loop tick.

        Deferring prevents an AttributeError that occurs when the signal is emitted
        synchronously inside process_sku_scan(), because showing a QMessageBox here
        (and the user clicking Yes → end_session() → self.logic = None) would corrupt
        the caller's stack frame that still holds references to self.logic.
        """
        QTimer.singleShot(0, self._show_all_complete_dialog)

    def _show_all_complete_dialog(self):
        """Show the 'all orders packed/processed' prompt after the current event loop cycle."""
        if not self.logic:
            return
        skipped_count = len(self.logic.session_packing_state.get('skipped_orders', []))
        if skipped_count:
            msg = (
                f"All processable orders have been packed!\n"
                f"{skipped_count} order(s) were skipped.\n\n"
                f"End session now?"
            )
        else:
            msg = "All orders have been packed!\nEnd session now?"
        reply = QMessageBox.question(
            self,
            "Session Complete",
            msg,
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.Yes,
        )
        if reply == QMessageBox.StandardButton.Yes:
            self.end_session()

    # REMOVED: open_restore_session_dialog() method (dead code)
    # This method was never called. Functionality replaced by Session Browser's
    # Active/Completed tabs which provide a better UX for session restoration

    def open_shopify_session(self):
        """
        Open Shopify session and select packing list to work on.

        Phase 1.8 Enhanced workflow:
        1. Use SessionSelectorDialog to browse Shopify sessions
        2. Automatically scan packing_lists/ folder
        3. User can select specific packing list or load entire session
        4. Create work directory: packing/{list_name}/ for selected lists

        If a client is already selected in the main menu, it will be
        pre-selected in the dialog (no need to select twice).
        """
        logger.info("Opening Shopify session selector")

        # Check if client is selected
        if not self.current_client_id:
            logger.warning("Attempted to open Shopify session without selecting client")
            self.client_combo.setStyleSheet("border: 2px solid red;")
            QMessageBox.warning(
                self,
                "No Client Selected",
                "Please select a client before opening a Shopify session!"
            )
            QTimer.singleShot(2000, lambda: self.client_combo.setStyleSheet(""))
            return

        # Check if session already active
        if self.session_manager and self.session_manager.is_active():
            logger.warning("Attempted to open Shopify session while one is already active")
            QMessageBox.warning(
                self,
                "Session Active",
                "A session is already active. Please end it first."
            )
            return

        # Step 1: Use SessionSelectorDialog to select session and packing list
        # Pass pre-selected client from main menu to avoid double selection
        selector_dialog = SessionSelectorDialog(
            profile_manager=self.profile_manager,
            pre_selected_client=self.current_client_id,
            parent=self
        )

        if not selector_dialog.exec():
            logger.info("Shopify session selection cancelled")
            return

        session_path = selector_dialog.get_selected_session()
        packing_list_path = selector_dialog.get_selected_packing_list()

        if not session_path:
            logger.warning("No session selected")
            return

        logger.info(f"Selected Shopify session: {session_path}")

        # Step 2: Determine loading mode
        if packing_list_path:
            # User selected a specific packing list
            selected_name = packing_list_path.stem
            logger.info(f"Selected packing list: {selected_name}")
            load_mode = "packing_list"
        else:
            # User wants to load entire session (analysis_data.json)
            logger.info("Loading entire session (analysis_data.json)")
            selected_name = "full_session"
            load_mode = "full_session"

        # Step 3: Create work directory structure
        try:
            # Create SessionManager for this client (not initialized yet)
            if not self.session_manager:
                self.session_manager = SessionManager(
                    client_id=self.current_client_id,
                    profile_manager=self.profile_manager,
                    lock_manager=self.lock_manager,
                    worker_id=self.current_worker_id,
                    worker_name=self.current_worker_name
                )

            # Determine packing list name and work directory
            packing_list_name = selected_name if load_mode == "packing_list" else "full_session"
            work_dir = self.session_manager.get_packing_work_dir(
                session_path=str(session_path),
                packing_list_name=packing_list_name
            )

            logger.info(f"Work directory created via SessionManager: {work_dir}")

            # Handle based on mode
            if load_mode == "packing_list":
                # Use unified session start method for packing list mode
                success = self.start_shopify_packing_session(
                    packing_list_path=packing_list_path,
                    work_dir=work_dir,
                    session_path=session_path,
                    client_id=self.current_client_id,
                    packing_list_name=packing_list_name
                )

                if not success:
                    # Error already shown by unified method
                    return

                # Get order count for success message
                order_count = self.packing_data.get('total_orders', 0)
                list_name = self.packing_data.get('list_name', packing_list_name)

                QMessageBox.information(
                    self,
                    "Session Loaded",
                    f"Session: {session_path.name}\n"
                    f"Packing List: {selected_name}\n"
                    f"Orders: {order_count}\n\n"
                    f"Work directory:\n{work_dir}"
                )

            else:
                # Handle full_session mode (load entire session from analysis_data.json)
                # This mode uses different loading logic and is kept separate
                logger.info(f"Loading full session from: {session_path}")

                # Store current session info
                self.current_session_path = str(session_path)
                self.current_packing_list = selected_name
                self.current_work_dir = str(work_dir)

                # Acquire lock
                success, error_msg, _ = self.lock_manager.acquire_lock(
                    self.current_client_id,
                    work_dir,
                    worker_id=self.current_worker_id,
                    worker_name=self.current_worker_name
                )

                if not success:
                    if error_msg and "stale" in error_msg.lower():
                        reply = QMessageBox.question(
                            self, "Stale Lock Detected",
                            f"{error_msg}\n\nForce-release lock and continue?",
                            QMessageBox.Yes | QMessageBox.No
                        )
                        if reply == QMessageBox.Yes:
                            self.lock_manager.force_release_lock(work_dir)
                            success, error_msg, _ = self.lock_manager.acquire_lock(
                                self.current_client_id,
                                work_dir,
                                worker_id=self.current_worker_id,
                                worker_name=self.current_worker_name
                            )
                            if not success:
                                QMessageBox.warning(self, "Lock Failed", f"Failed to acquire lock: {error_msg}")
                                return
                        else:
                            return
                    else:
                        QMessageBox.warning(self, "Session Locked", error_msg or "Session is locked by another user")
                        return

                logger.info(f"Lock acquired for work directory: {work_dir}")

                # Create PackerLogic instance
                self.logic = PackerLogic(
                    client_id=self.current_client_id,
                    profile_manager=self.profile_manager,
                    work_dir=str(work_dir)
                )

                # Connect signals
                self.logic.item_packed.connect(self._on_item_packed)
                self.logic.all_orders_complete.connect(self._on_all_orders_complete)

                # Start heartbeat timer
                self._start_heartbeat_timer()

                # Load entire session (analysis_data.json)
                order_count, analyzed_at = self.logic.load_from_shopify_analysis(session_path)
                logger.info(f"Loaded full session: {order_count} orders (analyzed at {analyzed_at})")

                # Load analysis_data.json for UI display
                try:
                    analysis_file = session_path / "analysis" / "analysis_data.json"
                    with open(analysis_file, 'r', encoding='utf-8') as f:
                        self.packing_data = json.load(f)
                except Exception as e:
                    logger.warning(f"Could not load analysis data: {e}")
                    self.packing_data = {
                        'analyzed_at': analyzed_at,
                        'total_orders': order_count,
                        'orders': []
                    }

                # Update session metadata
                if hasattr(self.session_manager, 'update_session_metadata'):
                    try:
                        self.session_manager.update_session_metadata(
                            self.current_session_path,
                            self.current_packing_list,
                            'in_progress'
                        )
                    except Exception as e:
                        logger.warning(f"Could not update session metadata: {e}")

                # Setup order table
                self.setup_order_table()

                # Update UI with loaded data
                self.status_label.setText(
                    f"Loaded: {session_path.name} / {selected_name}\n"
                    f"Orders: {order_count}\n"
                    f"Ready to start packing"
                )

                # Enable packing UI
                self.enable_packing_mode()

                QMessageBox.information(
                    self,
                    "Session Loaded",
                    f"Session: {session_path.name}\n"
                    f"Mode: {selected_name}\n"
                    f"Orders: {order_count}\n\n"
                    f"Work directory:\n{work_dir}"
                )

        except Exception as e:
            # Only handle exceptions from full_session mode
            # (packing_list mode errors are handled by unified method)
            logger.error(f"Failed to load session: {e}", exc_info=True)
            self._cleanup_failed_session_start()

            # Determine error message based on exception type
            if isinstance(e, FileNotFoundError):
                title = "File Not Found"
                message = f"Session file not found:\n{str(e)}"
            elif isinstance(e, json.JSONDecodeError):
                title = "Invalid JSON"
                message = f"Session contains invalid JSON:\n{str(e)}"
            elif isinstance(e, (KeyError, ValueError)):
                title = "Invalid Data"
                message = f"Session data validation failed:\n{str(e)}"
            elif isinstance(e, RuntimeError):
                title = "Session Load Failed"
                message = f"Failed to load session:\n{str(e)}"
            else:
                title = "Error"
                message = f"Unexpected error loading session:\n{str(e)}"

            QMessageBox.critical(self, title, message)

    def enable_packing_mode(self):
        """
        Enable packing UI after session data is loaded.

        This method:
        - Disables session start buttons
        - Enables packing operation buttons
        - Updates status message and toolbar
        - Prepares UI for packing operations
        """
        logger.info("Enabling packing mode UI")

        # Enable packing operation buttons
        self.packer_mode_button.setEnabled(True)

        # Enable toolbar end button
        if hasattr(self, 'toolbar_end_btn'):
            self.toolbar_end_btn.setEnabled(True)

        # Update session info label in toolbar
        if hasattr(self, 'session_info_label') and hasattr(self, 'current_packing_list'):
            session_name = self.current_packing_list if self.current_packing_list else "Active Session"
            self.session_info_label.setText(f"Session: {session_name}")

        # Update status
        if hasattr(self, 'packing_data') and self.packing_data:
            self.status_label.setText(
                f"Ready to pack: {self.current_packing_list}\n"
                f"Orders: {self.packing_data.get('total_orders', 0)}"
            )
        else:
            self.status_label.setText("Ready to pack")

        logger.info("Packing mode UI enabled successfully")

    def open_session_browser(self):
        """Open Session Browser dialog."""
        logger.info("Opening Session Browser")

        # Create dialog
        browser_dialog = QDialog(self)
        browser_dialog.setWindowTitle("Session Browser")
        browser_dialog.setMinimumSize(1000, 700)
        browser_dialog.setModal(False)  # Non-modal - can keep open while working

        layout = QVBoxLayout(browser_dialog)

        # Create Session Browser widget
        browser = SessionBrowserWidget(
            profile_manager=self.profile_manager,
            session_manager=self.session_manager,
            session_lock_manager=self.lock_manager,
            session_history_manager=self.session_history_manager,
            worker_manager=self.worker_manager,
            registry_manager=self.registry_manager,
            parent=browser_dialog
        )

        # Connect signals
        browser.resume_session_requested.connect(
            lambda info: self._handle_resume_session_from_browser(browser_dialog, info)
        )
        browser.start_packing_requested.connect(
            lambda info: self._handle_start_packing_from_browser(browser_dialog, info)
        )

        layout.addWidget(browser)

        # Show dialog
        browser_dialog.exec()

    def _handle_resume_session_from_browser(self, dialog, session_info: dict):
        """
        Handle resume request from Session Browser.

        Args:
            dialog: Session Browser dialog to close
            session_info: Dict with session_path, client_id, packing_list_name, work_dir
        """
        logger.info(f"Resuming session from browser: {session_info.get('session_id', 'Unknown')}")

        # Close browser dialog
        dialog.accept()

        # Extract info
        session_path = Path(session_info['session_path'])
        client_id = session_info['client_id']
        packing_list_name = session_info['packing_list_name']
        work_dir = Path(session_info['work_dir'])

        # Set current client if different
        if self.current_client_id != client_id:
            # Find client index in combo
            for i in range(self.client_combo.count()):
                if self.client_combo.itemData(i) == client_id:
                    self.client_combo.setCurrentIndex(i)
                    break

        # Check if session already active
        if self.session_manager and self.session_manager.is_active():
            logger.warning("Attempted to resume session while one is already active")
            QMessageBox.warning(
                self,
                "Session Active",
                "A session is already active. Please end it first."
            )
            return

        # Create SessionManager for this client if not exists
        if not self.session_manager or self.session_manager.client_id != client_id:
            self.session_manager = SessionManager(
                client_id=client_id,
                profile_manager=self.profile_manager,
                lock_manager=self.lock_manager,
                worker_id=self.current_worker_id,
                worker_name=self.current_worker_name
            )

        # Build packing list path
        packing_list_path = session_path / "packing_lists" / f"{packing_list_name}.json"

        # Use unified session start method
        success = self.start_shopify_packing_session(
            packing_list_path=packing_list_path,
            work_dir=work_dir,
            session_path=session_path,
            client_id=client_id,
            packing_list_name=packing_list_name
        )

        if success:
            # Get order count for success message
            order_count = self.packing_data.get('total_orders', 0) if hasattr(self, 'packing_data') else 0
            list_name = self.packing_data.get('list_name', packing_list_name) if hasattr(self, 'packing_data') else packing_list_name

            QMessageBox.information(
                self,
                "Session Resumed",
                f"Successfully resumed packing list: {list_name}\n"
                f"Orders: {order_count}\n\n"
                f"Continue packing from where you left off."
            )
            logger.info("Session resumed successfully from Session Browser")

    def _handle_start_packing_from_browser(self, dialog, packing_info: dict):
        """
        Handle start packing request from Session Browser Available tab.

        Args:
            dialog: Session Browser dialog to close
            packing_info: Dict with session_path, client_id, packing_list_name, list_file
        """
        logger.info(f"Starting packing session from browser: {packing_info.get('packing_list_name', 'Unknown')}")

        # Close browser dialog
        dialog.accept()

        # Extract info
        session_path = Path(packing_info['session_path'])
        client_id = packing_info['client_id']
        packing_list_name = packing_info['packing_list_name']
        packing_list_path = Path(packing_info['list_file'])

        # Set current client if different
        if self.current_client_id != client_id:
            # Find client index in combo
            for i in range(self.client_combo.count()):
                if self.client_combo.itemData(i) == client_id:
                    self.client_combo.setCurrentIndex(i)
                    break

        # Check if session already active
        if self.session_manager and self.session_manager.is_active():
            logger.warning("Attempted to start packing while session is already active")
            QMessageBox.warning(
                self,
                "Session Active",
                "A session is already active. Please end it first."
            )
            return

        # Create SessionManager for this client if not exists
        if not self.session_manager or self.session_manager.client_id != client_id:
            self.session_manager = SessionManager(
                client_id=client_id,
                profile_manager=self.profile_manager,
                lock_manager=self.lock_manager,
                worker_id=self.current_worker_id,
                worker_name=self.current_worker_name
            )

        # Create work directory via SessionManager
        work_dir = self.session_manager.get_packing_work_dir(
            session_path=str(session_path),
            packing_list_name=packing_list_name
        )

        logger.info(f"Work directory created: {work_dir}")

        # Use unified session start method
        success = self.start_shopify_packing_session(
            packing_list_path=packing_list_path,
            work_dir=work_dir,
            session_path=session_path,
            client_id=client_id,
            packing_list_name=packing_list_name
        )

        if success:
            # Get order count for success message
            order_count = self.packing_data.get('total_orders', 0) if hasattr(self, 'packing_data') else 0
            list_name = self.packing_data.get('list_name', packing_list_name) if hasattr(self, 'packing_data') else packing_list_name

            QMessageBox.information(
                self,
                "Session Loaded",
                f"Loaded packing list: {list_name}\n"
                f"Orders: {order_count}\n\n"
                f"Ready to start packing."
            )
            logger.info("Packing session started successfully from Session Browser")

    def _handle_session_locked_error(self, error: SessionLockedError):
        """
        Handle when a session is actively locked by another process.

        Shows a dialog informing the user that the session is currently in use.

        Args:
            error: SessionLockedError with lock information
        """
        lock_info = error.lock_info
        if not lock_info:
            QMessageBox.warning(
                self,
                "Session Locked",
                "This session is currently locked by another process.\n\n"
                "Please wait or choose a different session."
            )
            return

        locked_by = lock_info.get('locked_by', 'Unknown PC')
        user_name = lock_info.get('user_name', 'Unknown user')
        lock_time = lock_info.get('lock_time', 'Unknown time')

        # Format time nicely
        try:
            from datetime import datetime
            lock_dt = datetime.fromisoformat(lock_time)
            lock_time_formatted = lock_dt.strftime('%d.%m.%Y %H:%M')
        except:
            lock_time_formatted = lock_time

        msg = QMessageBox(self)
        msg.setIcon(QMessageBox.Warning)
        msg.setWindowTitle("Session Already in Use")
        msg.setText("This session is currently active on another computer.")
        msg.setInformativeText(
            f"<b>User:</b> {user_name}<br>"
            f"<b>Computer:</b> {locked_by}<br>"
            f"<b>Started:</b> {lock_time_formatted}<br><br>"
            "Please wait for the user to finish, or choose another session."
        )
        msg.setStandardButtons(QMessageBox.Ok)
        msg.exec()

    def _handle_stale_lock_error(self, error: StaleLockError, file_path: str, restore_dir: str):
        """
        Handle when a session has a stale lock (possible crash).

        Shows a dialog allowing the user to force-release the lock and open the session.

        Args:
            error: StaleLockError with lock information
            file_path: Path to the packing list file
            restore_dir: Directory of the session to restore
        """
        lock_info = error.lock_info
        if not lock_info:
            QMessageBox.warning(self, "Stale Lock", "Session has an invalid lock file.")
            return

        locked_by = lock_info.get('locked_by', 'Unknown PC')
        user_name = lock_info.get('user_name', 'Unknown user')
        heartbeat = lock_info.get('heartbeat', 'Unknown')
        stale_minutes = error.stale_minutes

        # Format time nicely
        try:
            from datetime import datetime
            heartbeat_dt = datetime.fromisoformat(heartbeat)
            heartbeat_formatted = heartbeat_dt.strftime('%d.%m.%Y %H:%M')
        except:
            heartbeat_formatted = heartbeat

        msg = QMessageBox(self)
        msg.setIcon(QMessageBox.Warning)
        msg.setWindowTitle("Stale Session Lock Detected")
        msg.setText("This session has a stale lock - the application may have crashed.")
        msg.setInformativeText(
            f"<b>Original user:</b> {user_name}<br>"
            f"<b>Computer:</b> {locked_by}<br>"
            f"<b>Last heartbeat:</b> {heartbeat_formatted}<br>"
            f"<b>No response for:</b> {stale_minutes} minutes<br><br>"
            "The application may have crashed on that PC.<br><br>"
            "<b>Do you want to force-release the lock and open this session?</b>"
        )
        msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        msg.setDefaultButton(QMessageBox.Yes)

        reply = msg.exec()

        if reply == QMessageBox.Yes:
            # Force release the lock
            logger.info(f"User chose to force-release stale lock for session {restore_dir}")
            try:
                success = self.lock_manager.force_release_lock(Path(restore_dir))
                if success:
                    logger.info("Stale lock force-released successfully")
                    # Retry opening the session
                    QTimer.singleShot(100, lambda: self.start_session(file_path=file_path, restore_dir=restore_dir))
                else:
                    logger.error("Failed to force-release lock")
                    QMessageBox.critical(
                        self,
                        "Error",
                        "Failed to release the lock. Please try again or contact support."
                    )
            except Exception as e:
                logger.error(f"Error force-releasing lock: {e}", exc_info=True)
                QMessageBox.critical(
                    self,
                    "Error",
                    f"Failed to release lock:\n\n{e}"
                )
        else:
            logger.info("User cancelled force-release of stale lock")
            self.status_label.setText("Session opening cancelled.")


# ==============================================================================
# NEW CLIENT DIALOG
# ==============================================================================

class NewClientDialog(QDialog):
    """Dialog for creating a new client profile."""

    def __init__(self, profile_manager: ProfileManager, parent=None):
        """
        Initialize the new client dialog.

        Args:
            profile_manager: ProfileManager instance for validation and creation
            parent: Parent widget
        """
        super().__init__(parent)
        self.profile_manager = profile_manager
        self.client_id = None
        self.client_name = None

        self.setWindowTitle("Create New Client Profile")
        self.setMinimumWidth(400)

        layout = QFormLayout(self)

        # Instructions
        instructions = QLabel(
            "Create a new client profile. The client ID should be short (1-10 characters) "
            "and will be used to organize sessions and settings."
        )
        instructions.setWordWrap(True)
        layout.addRow(instructions)

        # Client ID input
        self.client_id_input = QLineEdit()
        self.client_id_input.setPlaceholderText("e.g., M, R, CLIENT_A")
        self.client_id_input.setMaxLength(10)
        self.client_id_input.textChanged.connect(self.on_id_changed)
        layout.addRow("Client ID (short):", self.client_id_input)

        # Validation label
        self.validation_label = QLabel("")
        self.validation_label.setStyleSheet("color: red;")
        self.validation_label.setWordWrap(True)
        layout.addRow("", self.validation_label)

        # Client name input
        self.client_name_input = QLineEdit()
        self.client_name_input.setPlaceholderText("e.g., M Cosmetics, R Fashion")
        layout.addRow("Client Full Name:", self.client_name_input)

        # Buttons
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        self.ok_button = button_box.button(QDialogButtonBox.Ok)
        self.ok_button.setEnabled(False)
        button_box.accepted.connect(self.accept_dialog)
        button_box.rejected.connect(self.reject)
        layout.addRow(button_box)

    def on_id_changed(self, text: str):
        """
        Validate client ID as user types.

        Args:
            text: Current text in client ID input
        """
        if not text:
            self.validation_label.setText("")
            self.ok_button.setEnabled(False)
            return

        # Convert to uppercase
        text_upper = text.upper()
        if text != text_upper:
            # Update to uppercase
            self.client_id_input.blockSignals(True)
            cursor_pos = self.client_id_input.cursorPosition()
            self.client_id_input.setText(text_upper)
            self.client_id_input.setCursorPosition(cursor_pos)
            self.client_id_input.blockSignals(False)
            text = text_upper

        # Validate
        is_valid, error_msg = self.profile_manager.validate_client_id(text)

        if not is_valid:
            self.validation_label.setText(f"⚠ {error_msg}")
            self.validation_label.setStyleSheet("color: red;")
            self.ok_button.setEnabled(False)
            return

        # Check if already exists
        if self.profile_manager.client_exists(text):
            self.validation_label.setText(f"⚠ Client '{text}' already exists!")
            self.validation_label.setStyleSheet("color: red;")
            self.ok_button.setEnabled(False)
            return

        # Valid
        self.validation_label.setText("✓ Valid client ID")
        self.validation_label.setStyleSheet("color: green;")
        self.ok_button.setEnabled(bool(self.client_name_input.text().strip()))

    def accept_dialog(self):
        """Handle OK button click."""
        client_id = self.client_id_input.text().strip().upper()
        client_name = self.client_name_input.text().strip()

        if not client_name:
            QMessageBox.warning(self, "Invalid Input", "Please enter a client name.")
            return

        logger.info(f"Creating new client: {client_id} - {client_name}")

        try:
            success = self.profile_manager.create_client_profile(client_id, client_name)

            if success:
                self.client_id = client_id
                self.client_name = client_name
                self.accept()
            else:
                QMessageBox.warning(
                    self,
                    "Error",
                    f"Client '{client_id}' already exists!"
                )

        except ValidationError as e:
            logger.error(f"Validation error creating client: {e}")
            QMessageBox.warning(self, "Validation Error", str(e))

        except Exception as e:
            logger.error(f"Error creating client: {e}", exc_info=True)
            QMessageBox.critical(
                self,
                "Error",
                f"Failed to create client profile:\n\n{e}"
            )


# ==============================================================================
# HELPER FUNCTIONS
# ==============================================================================

def restore_session(window: MainWindow):
    """
    Checks for and offers to restore an incomplete session on startup.

    Args:
        window (MainWindow): The main application window instance.
    """
    latest_dir = find_latest_session_dir()
    if not latest_dir:
        return

    session_info_path = os.path.join(latest_dir, "session_info.json")
    state_path = os.path.join(latest_dir, "packing_state.json")

    msg_box = QMessageBox()
    msg_box.setIcon(QMessageBox.Question)
    msg_box.setText("An incomplete session was found.")
    msg_box.setInformativeText(f"Would you like to restore session '{os.path.basename(latest_dir)}'?")
    msg_box.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
    msg_box.setDefaultButton(QMessageBox.Yes)
    reply = msg_box.exec()

    if reply == QMessageBox.Yes:
        try:
            with open(session_info_path, 'r') as f:
                session_info = json.load(f)
            packing_list_path = session_info.get('packing_list_path')
            if packing_list_path and os.path.exists(packing_list_path):
                window.start_session(file_path=packing_list_path, restore_dir=latest_dir)
            else:
                QMessageBox.critical(window, "Error", "Could not restore session. The original packing list file was not found.")
        except Exception as e:
            QMessageBox.critical(window, "Error", f"An error occurred during session restoration: {e}")
    else:
        try:
            if os.path.exists(session_info_path):
                os.rename(session_info_path, session_info_path + ".ignored")
            if os.path.exists(state_path):
                os.rename(state_path, state_path + ".ignored")
        except OSError as e:
            QMessageBox.warning(window, "Cleanup Warning", f"Could not ignore all session files: {e}")


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Packer's Assistant")
    parser.add_argument(
        '--config',
        default=DEFAULT_CONFIG_PATH,
        metavar='PATH',
        help=f'Path to config file (default: {DEFAULT_CONFIG_PATH}). '
             'Use config.dev.ini for local development / mock server.'
    )
    args = parser.parse_args()

    app = QApplication(sys.argv)

    # Load saved theme (dark by default)
    load_saved_theme(app)

    window = MainWindow(config_path=args.config)

    # Check for abandoned sessions before showing the main window
    restore_session(window)

    window.show()
    sys.exit(app.exec())
